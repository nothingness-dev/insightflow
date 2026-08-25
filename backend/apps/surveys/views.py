import csv
import io
from collections import defaultdict
from django.http import HttpResponse
from django.shortcuts import get_object_or_404
from django.utils import timezone
from django.db import IntegrityError, transaction
from django.db.models import Count, Max, Q, F
from rest_framework import status, generics
from rest_framework.views import APIView
from rest_framework.response import Response
from rest_framework.permissions import IsAuthenticated, AllowAny
from apps.accounts.permissions import IsAdminUser, IsEmployeeUser
from apps.accounts.throttles import AnonymousSurveyRateThrottle
from apps.core.export_security import sanitize_cell

from .models import Survey, SurveyQuestion, SurveyPerson, Rating, SurveyHashLink, AnonymousParticipation
from .serializers import (
    SurveySerializer, SurveyCreateUpdateSerializer, SurveyPersonSerializer,
    SurveyPersonPublicSerializer, SurveyPublicSerializer, RatingCreateSerializer,
    SurveyProgressDashboardSerializer, SurveyHashLinkSerializer, SurveyQuestionSerializer,
)
from .services import calculate_survey_results, duplicate_survey, calculate_survey_progress, validate_question_answers, effective_questions_for_person, required_question_pairs, completed_participants, completed_participants_for, completed_person_ids as participant_completed_person_ids, annotate_survey_list_stats, bulk_completed_response_counts
from apps.activity.models import ActivityActions
from apps.activity.services import log_activity
from .export_data import (
    build_export_dataset, build_pdf_comment_groups, score_grade, EXCEL_CELL_LIMIT,
)
import logging
from django.conf import settings
from django.core.cache import cache
from apps.core.cache import (
    key_dashboard, key_survey_results, key_employee_survey_list, key_hash_links,
    invalidate_dashboard, invalidate_survey_results,
    invalidate_all_employee_survey_lists, invalidate_employee_survey_list,
    invalidate_hash_links,
)

logger = logging.getLogger('apps')


def survey_results_cache_key(survey: Survey) -> str:
    ratings_state = survey.ratings.aggregate(
        count=Count('id'),
        latest=Max('created_at'),
    )
    people_state = survey.people.filter(is_active=True).aggregate(
        count=Count('id'),
        latest=Max('updated_at'),
    )
    questions_state = survey.questions.filter(is_active=True).aggregate(
        count=Count('id'),
        latest=Max('updated_at'),
    )
    latest_rating = ratings_state['latest']
    latest_person = people_state['latest']
    latest_question = questions_state['latest']
    signature = ':'.join([
        str(ratings_state['count'] or 0),
        latest_rating.isoformat() if latest_rating else 'none',
        str(people_state['count'] or 0),
        latest_person.isoformat() if latest_person else 'none',
        str(questions_state['count'] or 0),
        latest_question.isoformat() if latest_question else 'none',
        survey.updated_at.isoformat() if survey.updated_at else 'none',
    ])
    return f'{key_survey_results(survey.id)}:{signature}'

try:
    import openpyxl
    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False

try:
    from .pdf_report import build_survey_pdf
    HAS_PDF = True
    PDF_IMPORT_ERROR = None
except Exception as _pdf_import_exc:                                                             
    HAS_PDF = False
    PDF_IMPORT_ERROR = str(_pdf_import_exc)


from apps.activity.services import _client_ip as get_client_ip



class AdminSurveyListCreateView(generics.ListCreateAPIView):
    permission_classes = [IsAdminUser]
    queryset = Survey.objects.all()

    def get_serializer_class(self):
        if self.request.method == 'POST':
            return SurveyCreateUpdateSerializer
        return SurveySerializer

    def get_queryset(self):
        qs = super().get_queryset()
        search = self.request.query_params.get('search')
        if search:
            qs = (qs.filter(title__icontains=search) | qs.filter(question__icontains=search) | qs.filter(questions__text__icontains=search)).distinct()
        status_filter = self.request.query_params.get('status')
        if status_filter:
            qs = qs.filter(status=status_filter)
        return qs

    def list(self, request, *args, **kwargs):
        # Annotated counters + one bulk completion query keep the list at a
        # fixed query count instead of ~5 queries per survey row.
        queryset = annotate_survey_list_stats(self.filter_queryset(self.get_queryset())).select_related('created_by')
        page = self.paginate_queryset(queryset)
        surveys = page if page is not None else list(queryset)

        completed_counts = bulk_completed_response_counts([s.id for s in surveys])
        for survey in surveys:
            survey.bulk_total_responses = completed_counts.get(survey.id, 0)

        serializer = self.get_serializer(surveys, many=True)
        if page is not None:
            return self.get_paginated_response(serializer.data)
        return Response(serializer.data)

    def create(self, request, *args, **kwargs):
        serializer = SurveyCreateUpdateSerializer(data=request.data)
        serializer.is_valid(raise_exception=True)
        survey = serializer.save(created_by=request.user)
        logger.info(f"Admin {request.user.username} created survey: {survey.title}")
        log_activity(
            ActivityActions.SURVEY_CREATE,
            request=request,
            description=f'ایجاد نظرسنجی «{survey.title}»',
            target_type='survey',
            target_id=survey.id,
            target_repr=survey.title,
        )
        active_questions = survey.questions.filter(is_active=True).count()
        if active_questions:
            log_activity(
                ActivityActions.QUESTION_ADD,
                request=request,
                description=f'افزودن {active_questions} سوال به نظرسنجی «{survey.title}»',
                target_type='survey',
                target_id=survey.id,
                target_repr=survey.title,
                metadata={'questions_added': active_questions},
            )
        return Response(SurveySerializer(survey, context={'request': request}).data, status=status.HTTP_201_CREATED)


class AdminSurveyDetailView(generics.RetrieveUpdateDestroyAPIView):
    permission_classes = [IsAdminUser]
    queryset = Survey.objects.all()

    def get_serializer_class(self):
        if self.request.method in ['PUT', 'PATCH']:
            return SurveyCreateUpdateSerializer
        return SurveySerializer

    def update(self, request, *args, **kwargs):
        kwargs['partial'] = True
        instance = self.get_object()
        if instance.status == Survey.STATUS_CLOSED:
            return Response({'detail': 'نظرسنجی بسته شده قابل ویرایش نیست.'}, status=status.HTTP_400_BAD_REQUEST)
        if instance.status != Survey.STATUS_DRAFT:
            data = request.data.copy() if hasattr(request.data, 'copy') else dict(request.data)
            data.pop('questions', None)
            request._full_data = data

        before_texts = {q.id: q.text for q in instance.questions.all()}
        before_active = set(instance.questions.filter(is_active=True).values_list('id', flat=True))

        response = super().update(request, *args, **kwargs)

        instance.refresh_from_db()
        after_texts = {q.id: q.text for q in instance.questions.all()}
        after_active = set(instance.questions.filter(is_active=True).values_list('id', flat=True))
        added = len([qid for qid in after_texts if qid not in before_texts])
        deactivated = len(before_active - after_active)
        edited = len([
            qid for qid in (before_active & after_active)
            if before_texts.get(qid) != after_texts.get(qid)
        ])

        log_activity(
            ActivityActions.SURVEY_EDIT,
            request=request,
            description=f'ویرایش نظرسنجی «{instance.title}»',
            target_type='survey',
            target_id=instance.id,
            target_repr=instance.title,
        )
        if added:
            log_activity(
                ActivityActions.QUESTION_ADD,
                request=request,
                description=f'افزودن {added} سوال به نظرسنجی «{instance.title}»',
                target_type='survey', target_id=instance.id, target_repr=instance.title,
                metadata={'questions_added': added},
            )
        if edited:
            log_activity(
                ActivityActions.QUESTION_EDIT,
                request=request,
                description=f'ویرایش {edited} سوال در نظرسنجی «{instance.title}»',
                target_type='survey', target_id=instance.id, target_repr=instance.title,
                metadata={'questions_edited': edited},
            )
        if deactivated:
            log_activity(
                ActivityActions.QUESTION_DELETE,
                request=request,
                description=f'حذف {deactivated} سوال از نظرسنجی «{instance.title}»',
                target_type='survey', target_id=instance.id, target_repr=instance.title,
                metadata={'questions_deleted': deactivated},
            )
        invalidate_dashboard()
        invalidate_survey_results(instance.id)
        return response

    def destroy(self, request, *args, **kwargs):
        instance = self.get_object()
        logger.info(f"Admin {request.user.username} deleted survey: {instance.title}")
        survey_id, survey_title = instance.id, instance.title
        response = super().destroy(request, *args, **kwargs)
        invalidate_dashboard()
        invalidate_survey_results(survey_id)
        invalidate_all_employee_survey_lists()
        log_activity(
            ActivityActions.SURVEY_DELETE,
            request=request,
            description=f'حذف نظرسنجی «{survey_title}»',
            target_type='survey',
            target_id=survey_id,
            target_repr=survey_title,
        )
        return response


class AdminSurveyProgressView(APIView):
    """Return participation progress for all surveys using bounded aggregate queries."""
    permission_classes = [IsAdminUser]

    def get(self, request):
        progress_data = calculate_survey_progress()
        serializer = SurveyProgressDashboardSerializer(progress_data)
        return Response(serializer.data)


class AdminSurveyDuplicateView(APIView):
    """Duplicate survey settings and people into a new draft without responses."""
    permission_classes = [IsAdminUser]

    def post(self, request, pk):
        try:
            source_survey = Survey.objects.get(pk=pk)
        except Survey.DoesNotExist:
            return Response({'detail': 'نظرسنجی یافت نشد.'}, status=status.HTTP_404_NOT_FOUND)

        duplicate = duplicate_survey(source_survey, request.user)
        logger.info(
            'Admin %s duplicated survey %s into survey %s',
            request.user.username,
            source_survey.id,
            duplicate.id,
        )
        log_activity(
            ActivityActions.SURVEY_DUPLICATE,
            request=request,
            description=f'تکثیر نظرسنجی «{source_survey.title}» به «{duplicate.title}»',
            target_type='survey',
            target_id=duplicate.id,
            target_repr=duplicate.title,
            metadata={'source_survey_id': source_survey.id},
        )
        return Response(
            SurveySerializer(duplicate, context={'request': request}).data,
            status=status.HTTP_201_CREATED,
        )


class AdminSurveyPublishView(APIView):
    permission_classes = [IsAdminUser]

    def post(self, request, pk):
        try:
            survey = Survey.objects.get(pk=pk)
        except Survey.DoesNotExist:
            return Response({'detail': 'نظرسنجی یافت نشد.'}, status=status.HTTP_404_NOT_FOUND)

        if survey.status != Survey.STATUS_DRAFT:
            return Response({'detail': 'فقط نظرسنجی‌های پیش‌نویس قابل انتشار هستند.'}, status=status.HTTP_400_BAD_REQUEST)

        if not survey.title:
            return Response({'detail': 'عنوان نظرسنجی الزامی است.'}, status=status.HTTP_400_BAD_REQUEST)

        if not survey.questions.filter(is_active=True).exists():
            return Response({'detail': 'حداقل یک سوال فعال باید به نظرسنجی اضافه شود.'}, status=status.HTTP_400_BAD_REQUEST)

        if not survey.people.filter(is_active=True).exists():
            return Response({'detail': 'حداقل یک فرد فعال باید به نظرسنجی اضافه شود.'}, status=status.HTTP_400_BAD_REQUEST)

        survey.status = Survey.STATUS_PUBLISHED
        survey.published_at = timezone.now()
        survey.save()
        invalidate_dashboard()
        invalidate_all_employee_survey_lists()
        logger.info(f"Admin {request.user.username} published survey: {survey.title}")
        log_activity(
            ActivityActions.SURVEY_PUBLISH,
            request=request,
            description=f'انتشار نظرسنجی «{survey.title}»',
            target_type='survey',
            target_id=survey.id,
            target_repr=survey.title,
        )
        return Response(SurveySerializer(survey, context={'request': request}).data)


class AdminSurveyCloseView(APIView):
    permission_classes = [IsAdminUser]

    def post(self, request, pk):
        try:
            survey = Survey.objects.get(pk=pk)
        except Survey.DoesNotExist:
            return Response({'detail': 'نظرسنجی یافت نشد.'}, status=status.HTTP_404_NOT_FOUND)

        if survey.status != Survey.STATUS_PUBLISHED:
            return Response({'detail': 'فقط نظرسنجی‌های منتشرشده قابل بستن هستند.'}, status=status.HTTP_400_BAD_REQUEST)

        survey.status = Survey.STATUS_CLOSED
        survey.closed_at = timezone.now()
        survey.save()
        invalidate_dashboard()
        invalidate_survey_results(survey.id)
        invalidate_all_employee_survey_lists()
        logger.info(f"Admin {request.user.username} closed survey: {survey.title}")
        log_activity(
            ActivityActions.SURVEY_CLOSE,
            request=request,
            description=f'بستن نظرسنجی «{survey.title}»',
            target_type='survey',
            target_id=survey.id,
            target_repr=survey.title,
        )
        return Response(SurveySerializer(survey, context={'request': request}).data)


class AdminSurveyResultsView(APIView):
    permission_classes = [IsAdminUser]

    def get(self, request, pk):
        try:
            survey = Survey.objects.get(pk=pk)
        except Survey.DoesNotExist:
            return Response({'detail': 'نظرسنجی یافت نشد.'}, status=status.HTTP_404_NOT_FOUND)

        ck = survey_results_cache_key(survey)
        cached = cache.get(ck)
        if cached is not None:
            return Response(cached)

        results = calculate_survey_results(survey, request)
        payload = {
            'survey': SurveySerializer(survey, context={'request': request}).data,
            'results': results,
        }
        cache.set(ck, payload, settings.CACHE_TTL_SURVEY_RESULTS)
        return Response(payload)


class AdminSurveyExportCSVView(APIView):
    """Export survey results as a UTF-8 CSV (BOM for Excel compatibility).

    The CSV is organized into clearly labeled sections (separated by blank
    rows and a "### عنوان بخش ###" marker row) so it reads sensibly when
    opened in Excel/Sheets/Numbers, not just when parsed by a script:
      1) Survey info header
      2) Per-person results matrix (scores, emoji, comment counts)
      3) Question-by-question analysis
      4) Score distribution (+ emoji distribution if present)
      5) Full text comments — one row per comment, with person/question context

    To stay scalable when a single question collects hundreds of comments,
    the per-question columns in section 2 carry only the comment COUNT;
    every comment's full text is listed in section 5.
    """
    permission_classes = [IsAdminUser]

    def get(self, request, pk):
        try:
            survey = Survey.objects.get(pk=pk)
        except Survey.DoesNotExist:
            return Response({'detail': 'نظرسنجی یافت نشد.'}, status=status.HTTP_404_NOT_FOUND)

        ds = build_export_dataset(survey, request)
        comments_map, summary = ds['comments_map'], ds['summary']
        questions_meta = ds['questions_meta']
        result_groups = ds['result_groups']

        response = HttpResponse(content_type='text/csv; charset=utf-8-sig')
        response['Content-Disposition'] = f'attachment; filename="survey_{pk}_results.csv"'
        log_activity(
            ActivityActions.EXPORT_CSV,
            request=request,
            description=f'خروجی CSV نتایج نظرسنجی «{survey.title}»',
            target_type='survey',
            target_id=survey.id,
            target_repr=survey.title,
            metadata={'export_format': 'csv'},
        )
        writer = csv.writer(response)

        def section_title(title):
            writer.writerow([])
            writer.writerow([f'### {title} ###'])

        # ---- 1) Survey info header -------------------------------------------------
        section_title('اطلاعات نظرسنجی')
        writer.writerow(['عنوان نظرسنجی', survey.title])
        writer.writerow(['تاریخ خروجی', timezone.localtime().strftime('%Y-%m-%d %H:%M')])
        writer.writerow(['تعداد افراد', summary['people']])
        writer.writerow(['تعداد سوالات', summary['questions']])
        writer.writerow(['تعداد رأی‌دهنده (تکمیل‌کننده)', summary['voters']])
        writer.writerow(['میانگین کلی', summary['overall_avg'] if summary['overall_avg'] is not None else '—'])
        writer.writerow(['مجموع نظرات متنی', summary['total_comments']])

        # ---- 2) Per-person results matrix ------------------------------------------
        # People with the full default question set and people with a custom/partial
        # question set are kept in clearly separated blocks (never mixed in the same
        # rows) - each custom person gets their own dedicated sub-section too, since
        # each one may have answered a different subset of questions.
        section_title('نتایج به تفکیک افراد')

        def _headers_for(qs):
            headers = ['رتبه', 'نام و نام خانوادگی', 'واحد سازمانی', 'سمت',
                       'میانگین کلی', 'کیفیت', 'تعداد رأی‌دهنده', 'تعداد پاسخ امتیازی']
            for q in qs:
                if q.has_score:
                    headers.append(f"میانگین: {q.text}")
                    headers.append(f"تعداد پاسخ: {q.text}")
                if q.has_emoji:
                    headers.append(f"امتیاز ایموجی: {q.text}")
                    headers.append(f"تعداد پاسخ ایموجی: {q.text}")
                if q.has_comment:
                    headers.append(f"تعداد نظرات: {q.text}")
            return headers

        def _row_for(r, qs):
            row = [
                r['rank'], r['full_name'], r['department'] or '', r['role_title'] or '',
                r['average_score'] if r['average_score'] is not None else '',
                score_grade(r['average_score']),
                r['votes_count'], r['scored_answers_count'],
            ]
            by_q = {item['question_id']: item for item in r.get('question_results', [])}
            for q in qs:
                item = by_q.get(q.id, {})
                if q.has_score:
                    row.append(item.get('average_score') if item.get('average_score') is not None else '')
                    row.append(item.get('responses_count') or 0)
                if q.has_emoji:
                    row.append(item.get('average_emoji_label') or '')
                    row.append(item.get('emoji_responses_count') or 0)
                if q.has_comment:
                    row.append(len(comments_map.get((r['person_id'], q.id), [])))
            return row

        # Each result group (shared + one per particular person) has its own
        # question set, so headers are rebuilt per group instead of reused.
        for group in result_groups:
            if not group['results']:
                continue
            group_questions = group['questions']
            writer.writerow([])
            writer.writerow([f'-- {group["title"]} --'])
            writer.writerow(_headers_for(group_questions))
            for r in group['results']:
                writer.writerow(_row_for(r, group_questions))

        # ---- 3) Question-by-question analysis ---------------------------------------
        def _write_questions_meta(meta_list):
            writer.writerow(['#', 'متن سوال', 'میانگین کل', 'کیفیت', 'تعداد پاسخ', 'امتیاز ایموجی', 'تعداد پاسخ ایموجی', 'تعداد نظرات متنی'])
            for idx, q in enumerate(meta_list, 1):
                writer.writerow([
                    idx, q['text'],
                    q['avg'] if q['avg'] is not None else ('متنی' if not q['has_score'] else ''),
                    score_grade(q['avg']) if q['has_score'] else '—',
                    q['responses'] if q['has_score'] else '—',
                    q['emoji_avg_label'] if q['has_emoji'] else '—',
                    q['emoji_responses'] if q['has_emoji'] else '—',
                    q['comments'] if q['has_comment'] else '—',
                ])

        section_title('تحلیل سوال‌به‌سوال')
        _write_questions_meta(questions_meta)

        # Particular persons get their own titled question-analysis sub-table,
        # never merged into the shared table above.
        for group in result_groups[1:]:
            if not group['questions_meta']:
                continue
            writer.writerow([])
            writer.writerow([f'-- {group["title"]} --'])
            _write_questions_meta(group['questions_meta'])

        # ---- 4) Distributions --------------------------------------------------------
        section_title('توزیع امتیازات')
        writer.writerow(['دسته', 'تعداد افراد'])
        for label, count, _color in summary['distribution']:
            writer.writerow([label, count])

        if summary['emoji_distribution']:
            section_title('توزیع امتیاز ایموجی')
            writer.writerow(['دسته', 'تعداد پاسخ'])
            for label, count, _color in summary['emoji_distribution']:
                writer.writerow([label, count])

        # ---- 5) Full text comments ----------------------------------------------------
        section_title(f"نظرات متنی (مجموع {summary['total_comments']})")
        if ds['comments_flat']:
            writer.writerow(['#', 'نام فرد ارزیابی‌شده', 'واحد سازمانی', 'سوال', 'نظر'])
            for idx, (person_name, dept, q_text, comment) in enumerate(ds['comments_flat'], 1):
                writer.writerow([idx, person_name, dept, q_text, sanitize_cell(comment)])
        else:
            writer.writerow(['در این نظرسنجی هیچ نظر متنی ثبت نشده است.'])

        for group in result_groups[1:]:
            if not group['comments_flat']:
                continue
            writer.writerow([])
            writer.writerow([f'-- {group["title"]} --'])
            writer.writerow(['#', 'نام فرد ارزیابی‌شده', 'واحد سازمانی', 'سوال', 'نظر'])
            for idx, (person_name, dept, q_text, comment) in enumerate(group['comments_flat'], 1):
                writer.writerow([idx, person_name, dept, q_text, sanitize_cell(comment)])

        return response


class AdminSurveyExportExcelView(APIView):
    """Export survey results as a styled multi-sheet Excel workbook.

    Visual language matches the PDF report (indigo brand header, KPI summary,
    score distribution, colour-graded scores). Every comment is preserved on a
    dedicated sheet — one row per comment — which scales to hundreds of
    comments per question without bloating the grid sheets (those carry counts).
    """
    permission_classes = [IsAdminUser]

    def get(self, request, pk):
        if not HAS_OPENPYXL:
            return Response({'detail': 'خروجی اکسل در دسترس نیست.'}, status=status.HTTP_501_NOT_IMPLEMENTED)

        try:
            survey = Survey.objects.get(pk=pk)
        except Survey.DoesNotExist:
            return Response({'detail': 'نظرسنجی یافت نشد.'}, status=status.HTTP_404_NOT_FOUND)

        from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
        from openpyxl.utils import get_column_letter

        ds = build_export_dataset(survey, request)
        comments_map, summary = ds['comments_map'], ds['summary']
        questions_meta = ds['questions_meta']
        result_groups = ds['result_groups']

        BRAND_FILL    = PatternFill('solid', fgColor='4F46E5')
        BRAND2_FILL   = PatternFill('solid', fgColor='7C3AED')           
        HEADER_FILL   = PatternFill('solid', fgColor='1E293B')              
        HEADER_FONT   = Font(name='Calibri', bold=True, color='FFFFFF', size=10)
        SUBHEAD_FILL  = PatternFill('solid', fgColor='334155')
        SUBHEAD_FONT  = Font(bold=True, color='FFFFFF', size=10)
        BRAND_FONT    = Font(bold=True, color='FFFFFF', size=15)
        SUBTITLE_FONT = Font(color='E0E7FF', size=10)
        TITLE_FONT    = Font(bold=True, size=13, color='1E293B')
        MUTED_FONT    = Font(color='64748B', size=9)
        CENTER        = Alignment(horizontal='center', vertical='center', wrap_text=True)
        RIGHT         = Alignment(horizontal='right', vertical='center', wrap_text=True)
        THIN          = Side(style='thin', color='E2E8F0')
        BORDER        = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

        def score_fill(score):
            if score is None: return PatternFill('solid', fgColor='F8FAFC')
            if score < 4:     return PatternFill('solid', fgColor='FEF2F2')
            if score < 7:     return PatternFill('solid', fgColor='FFFBEB')
            return             PatternFill('solid', fgColor='F0FDF4')

        def score_font(score):
            if score is None: return Font(color='94A3B8', bold=True, size=10)
            if score < 4:     return Font(color='EF4444', bold=True, size=10)
            if score < 7:     return Font(color='F59E0B', bold=True, size=10)
            return             Font(color='10B981', bold=True, size=10)

        EMOJI_FILL_COLORS = {
            Rating.EMOJI_BAD: 'FEF2F2',
            Rating.EMOJI_AVERAGE: 'FFFBEB',
            Rating.EMOJI_GOOD: 'F7FEE7',
            Rating.EMOJI_EXCELLENT: 'F0FDF4',
        }
        EMOJI_FONT_COLORS = {
            Rating.EMOJI_BAD: 'EF4444',
            Rating.EMOJI_AVERAGE: 'F59E0B',
            Rating.EMOJI_GOOD: '84CC16',
            Rating.EMOJI_EXCELLENT: '10B981',
        }

        def emoji_fill(emoji_key):
            return PatternFill('solid', fgColor=EMOJI_FILL_COLORS.get(emoji_key, 'F8FAFC'))

        def emoji_font(emoji_key):
            return Font(color=EMOJI_FONT_COLORS.get(emoji_key, '94A3B8'), bold=True, size=10)

        def style_header_row(ws, row_num, fill, font, height=22):
            ws.row_dimensions[row_num].height = height
            for cell in ws[row_num]:
                cell.fill = fill
                cell.font = font
                cell.alignment = CENTER
                cell.border = BORDER

        wb = openpyxl.Workbook()

        ws0 = wb.active
        ws0.title = 'خلاصه'
        ws0.sheet_view.rightToLeft = True
        ws0.sheet_view.showGridLines = False

        ws0.merge_cells('A1:D2')
        ws0['A1'] = f'نتایج نظرسنجی: {survey.title}'
        ws0['A1'].fill = BRAND_FILL
        ws0['A1'].font = BRAND_FONT
        ws0['A1'].alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        for r_ in (1, 2):
            for c_ in range(1, 5):
                ws0.cell(row=r_, column=c_).fill = BRAND_FILL
        ws0.row_dimensions[1].height = 26
        ws0.row_dimensions[2].height = 14
        ws0.merge_cells('A3:D3')
        ws0['A3'] = 'گزارش تحلیلی نتایج نظرسنجی'
        ws0['A3'].fill = BRAND2_FILL
        ws0['A3'].font = SUBTITLE_FONT
        ws0['A3'].alignment = Alignment(horizontal='center', vertical='center')
        for c_ in range(1, 5):
            ws0.cell(row=3, column=c_).fill = BRAND2_FILL
        ws0.row_dimensions[3].height = 18

        avg = summary['overall_avg']
        kpis = [
            ('میانگین کل', avg if avg is not None else '—'),
            ('سوالات فعال', summary['questions']),
            ('افراد ارزیابی‌شونده', summary['people']),
            ('رأی‌دهندگان کامل', summary['voters']),
        ]

        for i, (label, value) in enumerate(kpis, 1):
            vc = ws0.cell(row=5, column=i, value=value)
            vc.alignment = CENTER
            vc.font = Font(bold=True, size=16, color='4F46E5')
            vc.border = BORDER
            if i == 1 and isinstance(value, (int, float)):
                vc.font = score_font(avg); vc.fill = score_fill(avg)
            lc = ws0.cell(row=6, column=i, value=label)
            lc.alignment = CENTER
            lc.font = MUTED_FONT
            lc.border = BORDER
        ws0.row_dimensions[5].height = 30
        ws0.row_dimensions[6].height = 18

        ws0.cell(row=8, column=1, value='بهترین امتیاز').font = SUBHEAD_FONT
        ws0.cell(row=8, column=1).fill = SUBHEAD_FILL
        ws0.cell(row=8, column=2, value=summary['best'] if summary['best'] is not None else '—').alignment = CENTER
        ws0.cell(row=9, column=1, value='ضعیف‌ترین امتیاز').font = SUBHEAD_FONT
        ws0.cell(row=9, column=1).fill = SUBHEAD_FILL
        ws0.cell(row=9, column=2, value=summary['worst'] if summary['worst'] is not None else '—').alignment = CENTER

        ws0.merge_cells('A11:D11')
        ws0['A11'] = 'توزیع امتیازات'
        ws0['A11'].font = TITLE_FONT
        ws0['A11'].alignment = RIGHT
        dist_hdr = 12
        ws0.cell(row=dist_hdr, column=1, value='دسته')
        ws0.cell(row=dist_hdr, column=2, value='تعداد افراد')
        ws0.cell(row=dist_hdr, column=3, value='درصد')
        style_header_row(ws0, dist_hdr, HEADER_FILL, HEADER_FONT)
        dist_total = sum(c for _l, c, _col in summary['distribution']) or 1
        rr = dist_hdr + 1
        for label, count, color in summary['distribution']:
            ws0.cell(row=rr, column=1, value=label).alignment = RIGHT
            cc = ws0.cell(row=rr, column=2, value=count); cc.alignment = CENTER
            pc = ws0.cell(row=rr, column=3, value=f'{round(count / dist_total * 100)}٪'); pc.alignment = CENTER
            chip = ws0.cell(row=rr, column=1)
            chip.fill = PatternFill('solid', fgColor=color.lstrip('#').upper())
            chip.font = Font(color='FFFFFF', bold=True, size=10)
            for c_ in range(1, 4):
                ws0.cell(row=rr, column=c_).border = BORDER
            rr += 1
        ws0.column_dimensions['A'].width = 26
        ws0.column_dimensions['B'].width = 18
        ws0.column_dimensions['C'].width = 14
        ws0.column_dimensions['D'].width = 22

        if summary['emoji_distribution']:
            emoji_title_row = rr + 1
            ws0.merge_cells(start_row=emoji_title_row, start_column=1, end_row=emoji_title_row, end_column=4)
            ws0.cell(row=emoji_title_row, column=1, value='توزیع امتیاز ایموجی')
            ws0.cell(row=emoji_title_row, column=1).font = TITLE_FONT
            ws0.cell(row=emoji_title_row, column=1).alignment = RIGHT
            emoji_hdr = emoji_title_row + 1
            ws0.cell(row=emoji_hdr, column=1, value='دسته')
            ws0.cell(row=emoji_hdr, column=2, value='تعداد پاسخ')
            ws0.cell(row=emoji_hdr, column=3, value='درصد')
            style_header_row(ws0, emoji_hdr, HEADER_FILL, HEADER_FONT)
            emoji_total = sum(c for _l, c, _col in summary['emoji_distribution']) or 1
            er = emoji_hdr + 1
            for label, count, color in summary['emoji_distribution']:
                ws0.cell(row=er, column=1, value=label).alignment = RIGHT
                cc = ws0.cell(row=er, column=2, value=count); cc.alignment = CENTER
                pc = ws0.cell(row=er, column=3, value=f'{round(count / emoji_total * 100)}٪'); pc.alignment = CENTER
                chip = ws0.cell(row=er, column=1)
                chip.fill = PatternFill('solid', fgColor=color.lstrip('#').upper())
                chip.font = Font(color='FFFFFF', bold=True, size=10)
                for c_ in range(1, 4):
                    ws0.cell(row=er, column=c_).border = BORDER
                er += 1

        ws1 = wb.create_sheet('نتایج فردی')
        ws1.sheet_view.rightToLeft = True

        ws1.append([f'رتبه‌بندی افراد: {survey.title}'])
        base_cols = ['رتبه', 'نام و نام خانوادگی', 'واحد سازمانی', 'سمت',
                     'میانگین کلی', 'کیفیت', 'رأی‌دهنده', 'پاسخ امتیازی']
        max_ncols = len(base_cols)
        ws1.append([])

        # Each result group (shared + one per particular person) has its own
        # question set, so headers/columns are rebuilt per group instead of
        # being shared - a particular person's private questions never share
        # a column with the general comparison table.
        for group in result_groups:
            if not group['results']:
                continue
            q_headers = []
            for q in group['questions']:
                if q.has_score:
                    q_headers.append(f"میانگین\n{q.text}")
                    q_headers.append(f"تعداد پاسخ\n{q.text}")
                if q.has_emoji:
                    q_headers.append(f"امتیاز ایموجی\n{q.text}")
                    q_headers.append(f"تعداد پاسخ ایموجی\n{q.text}")
                if q.has_comment:
                    q_headers.append(f"تعداد نظرات\n{q.text}")
            group_ncols = len(base_cols) + len(q_headers)
            max_ncols = max(max_ncols, group_ncols)

            title_row = ws1.max_row + 1
            ws1.append([group['title']])
            ws1.merge_cells(start_row=title_row, start_column=1, end_row=title_row, end_column=group_ncols)
            gt_cell = ws1.cell(row=title_row, column=1)
            gt_cell.font = SUBHEAD_FONT
            gt_cell.fill = SUBHEAD_FILL
            gt_cell.alignment = RIGHT
            ws1.row_dimensions[title_row].height = 20

            header_row = ws1.max_row + 1
            ws1.append(base_cols + q_headers)
            style_header_row(ws1, header_row, HEADER_FILL, HEADER_FONT, height=40 if q_headers else 22)

            for r in group['results']:
                avg_v = r['average_score']
                row = [
                    r['rank'], r['full_name'], r['department'] or '', r['role_title'] or '',
                    round(avg_v, 2) if avg_v is not None else '',
                    score_grade(avg_v),
                    r['votes_count'], r.get('scored_answers_count') or 0,
                ]
                by_q = {item['question_id']: item for item in r.get('question_results', [])}
                for q in group['questions']:
                    item = by_q.get(q.id, {})
                    if q.has_score:
                        q_avg = item.get('average_score')
                        row.append(round(q_avg, 2) if q_avg is not None else '')
                        row.append(item.get('responses_count') or 0)
                    if q.has_emoji:
                        row.append(item.get('average_emoji_label') or '')
                        row.append(item.get('emoji_responses_count') or 0)
                    if q.has_comment:
                        row.append(len(comments_map.get((r['person_id'], q.id), [])))
                ws1.append(row)
                data_row = ws1.max_row
                ws1.row_dimensions[data_row].height = 18
                for col_idx, cell in enumerate(ws1[data_row], 1):
                    cell.border = BORDER
                    cell.alignment = CENTER if col_idx in (1, 5, 6, 7, 8) else RIGHT
                    if col_idx == 5:
                        cell.fill = score_fill(avg_v); cell.font = score_font(avg_v)
                    elif col_idx == 6:
                        cell.font = score_font(avg_v)

                offset = len(base_cols)
                for q in group['questions']:
                    if q.has_score:
                        q_avg_val = by_q.get(q.id, {}).get('average_score')
                        cell = ws1.cell(row=data_row, column=offset + 1)
                        cell.fill = score_fill(q_avg_val); cell.font = score_font(q_avg_val)
                        offset += 2
                    if q.has_emoji:
                        q_emoji_label = by_q.get(q.id, {}).get('average_emoji_label')
                        emoji_key = next((k for k, label in Rating.EMOJI_CHOICES if label == q_emoji_label), None)
                        cell = ws1.cell(row=data_row, column=offset + 1)
                        cell.fill = emoji_fill(emoji_key); cell.font = emoji_font(emoji_key)
                        offset += 2
                    if q.has_comment:
                        offset += 1

        ws1.merge_cells(start_row=1, start_column=1, end_row=1, end_column=max_ncols)
        ws1['A1'].font = TITLE_FONT
        ws1['A1'].alignment = RIGHT
        ws1.row_dimensions[1].height = 26
        ws1.freeze_panes = 'A4'
        for col, w in (('A', 8), ('B', 24), ('C', 18), ('D', 18), ('E', 12), ('F', 10), ('G', 10), ('H', 12)):
            ws1.column_dimensions[col].width = w
        for ci in range(len(base_cols) + 1, max_ncols + 1):
            ws1.column_dimensions[get_column_letter(ci)].width = 14

        ws2 = wb.create_sheet('تحلیل سوال‌ها')
        ws2.sheet_view.rightToLeft = True
        ws2.append([f'تحلیل سوال‌به‌سوال: {survey.title}'])
        ws2.merge_cells('A1:H1')
        ws2['A1'].font = TITLE_FONT
        ws2['A1'].alignment = RIGHT
        ws2.row_dimensions[1].height = 26
        ws2.append([])

        def _append_questions_meta_block(meta_list, title=None):
            if title is not None:
                title_row = ws2.max_row + 1
                ws2.append([title])
                ws2.merge_cells(start_row=title_row, start_column=1, end_row=title_row, end_column=8)
                tc = ws2.cell(row=title_row, column=1)
                tc.font = SUBHEAD_FONT
                tc.fill = SUBHEAD_FILL
                tc.alignment = RIGHT
                ws2.row_dimensions[title_row].height = 20
            header_row = ws2.max_row + 1
            ws2.append(['#', 'متن سوال', 'میانگین کل', 'کیفیت', 'تعداد پاسخ', 'امتیاز ایموجی', 'تعداد پاسخ ایموجی', 'تعداد نظرات متنی'])
            style_header_row(ws2, header_row, HEADER_FILL, HEADER_FONT)
            for idx, q in enumerate(meta_list, 1):
                q_avg = q['avg']
                q_emoji_label = q['emoji_avg_label']
                ws2.append([
                    idx, q['text'],
                    q_avg if q_avg is not None else 'متنی',
                    score_grade(q_avg) if q['has_score'] else '—',
                    q['responses'] if q['has_score'] else '—',
                    q_emoji_label if q['has_emoji'] else '—',
                    q['emoji_responses'] if q['has_emoji'] else '—',
                    q['comments'] if q['has_comment'] else '—',
                ])
                row_num = ws2.max_row
                ws2.row_dimensions[row_num].height = 20
                for col_idx, cell in enumerate(ws2[row_num], 1):
                    cell.border = BORDER
                    cell.alignment = CENTER if col_idx in (1, 3, 4, 5, 6, 7, 8) else RIGHT
                    if col_idx == 3 and isinstance(cell.value, (int, float)):
                        cell.fill = score_fill(q_avg); cell.font = score_font(q_avg)
                    elif col_idx == 4 and q['has_score']:
                        cell.font = score_font(q_avg)
                    elif col_idx == 6 and q['has_emoji']:
                        emoji_key = next((k for k, label in Rating.EMOJI_CHOICES if label == q_emoji_label), None)
                        cell.fill = emoji_fill(emoji_key); cell.font = emoji_font(emoji_key)

        _append_questions_meta_block(questions_meta)

        # Particular persons get their own titled question-analysis block,
        # never merged into the shared table above.
        for group in result_groups[1:]:
            if not group['questions_meta']:
                continue
            ws2.append([])
            _append_questions_meta_block(group['questions_meta'], title=group['title'])

        ws2.freeze_panes = 'A4'
        ws2.auto_filter.ref = 'A3:H3'
        for col, w in (('A', 6), ('B', 40), ('C', 12), ('D', 10), ('E', 12), ('F', 12), ('G', 14), ('H', 14)):
            ws2.column_dimensions[col].width = w

        any_comments = ds['comments_flat'] or any(g['comments_flat'] for g in result_groups[1:])
        if any_comments:
            ws3 = wb.create_sheet('نظرات متنی')
            ws3.sheet_view.rightToLeft = True
            ws3.append([f"نظرات متنی: {survey.title}  (مجموع {summary['total_comments']})"])
            ws3.merge_cells('A1:E1')
            ws3['A1'].font = TITLE_FONT
            ws3['A1'].alignment = RIGHT
            ws3.row_dimensions[1].height = 26
            ws3.append([])

            def _append_comments_block(comments_flat, title=None):
                if title is not None:
                    title_row = ws3.max_row + 1
                    ws3.append([title])
                    ws3.merge_cells(start_row=title_row, start_column=1, end_row=title_row, end_column=5)
                    tc = ws3.cell(row=title_row, column=1)
                    tc.font = SUBHEAD_FONT
                    tc.fill = SUBHEAD_FILL
                    tc.alignment = RIGHT
                    ws3.row_dimensions[title_row].height = 20
                header_row = ws3.max_row + 1
                ws3.append(['#', 'نام فرد ارزیابی‌شده', 'واحد سازمانی', 'سوال', 'نظر'])
                style_header_row(ws3, header_row, SUBHEAD_FILL, SUBHEAD_FONT)
                for i, (person_name, dept, q_text, comment) in enumerate(comments_flat, 1):
                    safe_comment = comment if len(comment) <= EXCEL_CELL_LIMIT else comment[:EXCEL_CELL_LIMIT] + '…'
                    ws3.append([i, person_name, dept, q_text, sanitize_cell(safe_comment)])
                    row_num = ws3.max_row
                    for col_idx, cell in enumerate(ws3[row_num], 1):
                        cell.border = BORDER
                        cell.alignment = CENTER if col_idx == 1 else RIGHT

            _append_comments_block(ds['comments_flat'])
            for group in result_groups[1:]:
                if not group['comments_flat']:
                    continue
                ws3.append([])
                _append_comments_block(group['comments_flat'], title=group['title'])

            ws3.freeze_panes = 'A4'
            ws3.auto_filter.ref = 'A3:E3'
            for col, w in (('A', 6), ('B', 22), ('C', 16), ('D', 32), ('E', 60)):
                ws3.column_dimensions[col].width = w

        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        response = HttpResponse(
            output.getvalue(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = f'attachment; filename="survey_{pk}_results.xlsx"'
        log_activity(
            ActivityActions.EXPORT_EXCEL,
            request=request,
            description=f'خروجی Excel نتایج نظرسنجی «{survey.title}»',
            target_type='survey',
            target_id=survey.id,
            target_repr=survey.title,
            metadata={'export_format': 'excel'},
        )
        return response


class AdminSurveyExportPDFView(APIView):
    """Export survey results as a beautiful, comprehensive RTL PDF report.

    The PDF is an executive summary: comments are grouped by question and
    capped (PDF_MAX_COMMENTS_PER_QUESTION / PDF_MAX_TOTAL_COMMENTS) so the
    document stays readable and bounded even when a question gathers hundreds
    of comments. Readers are pointed to the Excel/CSV export for the full list.
    """
    permission_classes = [IsAdminUser]

    def get(self, request, pk):
        if not HAS_PDF:
            return Response({'detail': 'خروجی PDF در دسترس نیست.'}, status=status.HTTP_501_NOT_IMPLEMENTED)

        try:
            survey = Survey.objects.get(pk=pk)
        except Survey.DoesNotExist:
            return Response({'detail': 'نظرسنجی یافت نشد.'}, status=status.HTTP_404_NOT_FOUND)

        try:
            ds = build_export_dataset(survey, request)
            comment_groups, truncated = build_pdf_comment_groups(ds)
            custom_sections = []
            for group in ds['result_groups'][1:]:
                group_comments, group_truncated = build_pdf_comment_groups(ds, group=group)
                custom_sections.append({
                    'title': group['title'],
                    'questions_meta': group['questions_meta'],
                    'comment_groups': group_comments,
                    'truncated': group_truncated,
                })
            pdf_bytes = build_survey_pdf(
                survey, ds['results'], ds['questions_meta'],
                comment_groups, ds['summary'], comments_truncated=truncated,
                custom_sections=custom_sections,
            )
        except Exception:

            logger.exception('PDF export failed for survey %s', pk)
            return Response(
                {'detail': 'خطا در تولید خروجی PDF. لطفاً دوباره تلاش کنید یا خروجی Excel/CSV را امتحان کنید.'},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR,
            )

        log_activity(
            ActivityActions.EXPORT_PDF,
            request=request,
            description=f'خروجی PDF نتایج نظرسنجی «{survey.title}»',
            target_type='survey',
            target_id=survey.id,
            target_repr=survey.title,
            metadata={'export_format': 'pdf'},
        )

        response = HttpResponse(pdf_bytes, content_type='application/pdf')
        response['Content-Disposition'] = f'attachment; filename="survey_{pk}_results.pdf"'
        return response



class AdminPersonListCreateView(generics.ListCreateAPIView):
    permission_classes = [IsAdminUser]
    serializer_class = SurveyPersonSerializer

    def get_queryset(self):
        survey_id = self.kwargs['survey_id']
        return SurveyPerson.objects.filter(survey_id=survey_id)

    def perform_create(self, serializer):
        survey_id = self.kwargs['survey_id']
        try:
            survey = Survey.objects.get(pk=survey_id)
        except Survey.DoesNotExist:
            from rest_framework.exceptions import NotFound
            raise NotFound('نظرسنجی یافت نشد.')
        if survey.status != Survey.STATUS_DRAFT:
            from rest_framework.exceptions import PermissionDenied
            raise PermissionDenied('افزودن فرد فقط در حالت پیش‌نویس امکان‌پذیر است.')
        person = serializer.save(survey=survey)
        log_activity(
            ActivityActions.PERSON_ADD,
            request=self.request,
            description=f'افزودن «{person.full_name}» به نظرسنجی «{survey.title}»',
            target_type='survey',
            target_id=survey.id,
            target_repr=survey.title,
            # The database primary key is an internal, globally increasing
            # identifier. It is not a survey-local person number and exposing
            # it here made normal sequence growth look like a data problem.
            metadata={'person_name': person.full_name},
        )
        invalidate_dashboard()
        invalidate_survey_results(survey_id)


class AdminPersonDetailView(generics.RetrieveUpdateDestroyAPIView):
    permission_classes = [IsAdminUser]
    serializer_class = SurveyPersonSerializer
    queryset = SurveyPerson.objects.all()

    def _check_draft(self, instance):
        if instance.survey.status != Survey.STATUS_DRAFT:
            from rest_framework.exceptions import PermissionDenied
            raise PermissionDenied('ویرایش یا حذف فرد فقط در حالت پیش‌نویس امکان‌پذیر است.')

    def update(self, request, *args, **kwargs):
        kwargs['partial'] = True
        instance = self.get_object()
        self._check_draft(instance)
        response = super().update(request, *args, **kwargs)
        log_activity(
            ActivityActions.PERSON_EDIT,
            request=request,
            description=f'ویرایش «{instance.full_name}» در نظرسنجی «{instance.survey.title}»',
            target_type='survey',
            target_id=instance.survey.id,
            target_repr=instance.survey.title,
            metadata={'person_id': instance.id, 'person_name': instance.full_name},
        )
        invalidate_dashboard()
        return response

    def destroy(self, request, *args, **kwargs):
        instance = self.get_object()
        self._check_draft(instance)
        survey, person_name, person_id = instance.survey, instance.full_name, instance.id
        response = super().destroy(request, *args, **kwargs)
        log_activity(ActivityActions.PERSON_DELETE, request=request,
            description=f'حذف «{person_name}» از نظرسنجی «{survey.title}»',
            target_type='survey', target_id=survey.id, target_repr=survey.title,
            metadata={'person_id': person_id, 'person_name': person_name})
        invalidate_dashboard()
        return response


class AdminPersonQuestionsView(APIView):
    permission_classes = [IsAdminUser]

    def put(self, request, pk):
        person = get_object_or_404(SurveyPerson.objects.select_related('survey'), pk=pk)
        if person.survey.status != Survey.STATUS_DRAFT:
            return Response({'detail': 'تخصیص سوال فقط در حالت پیش‌نویس قابل تغییر است.'}, status=403)

        use_default = request.data.get('use_default_questions')

        if use_default:
            # Revert this person to the survey-wide shared question set and
            # permanently delete their private questions - they belong only
            # to this person under the new model, so there is nothing to keep.
            person.custom_questions.all().delete()
            person.uses_default_questions = True
            person.save(update_fields=['uses_default_questions', 'updated_at'])
            invalidate_survey_results(person.survey_id)
            invalidate_dashboard()
            log_activity(ActivityActions.PERSON_EDIT, request=request,
                description=f'بازگشت «{person.full_name}» به سوال‌های پیش‌فرض',
                target_type='survey', target_id=person.survey_id, target_repr=person.survey.title,
                metadata={'person_id': person.id, 'custom_questions': False})
            return Response(SurveyPersonSerializer(person, context={'request': request}).data)

        questions_data = request.data.get('questions')
        if not isinstance(questions_data, list) or not questions_data:
            return Response({'questions': 'برای این فرد باید حداقل یک سوال اختصاصی تعریف شود.'}, status=400)

        serializer = SurveyQuestionSerializer(data=questions_data, many=True)
        serializer.is_valid(raise_exception=True)

        with transaction.atomic():
            person.custom_questions.all().delete()
            SurveyQuestion.objects.bulk_create([
                SurveyQuestion(
                    survey=person.survey,
                    person=person,
                    display_order=index,
                    is_active=True,
                    **{k: v for k, v in item.items() if k not in ('display_order', 'is_active')},
                )
                for index, item in enumerate(serializer.validated_data)
            ])
            person.uses_default_questions = False
            person.save(update_fields=['uses_default_questions', 'updated_at'])

        invalidate_survey_results(person.survey_id)
        invalidate_dashboard()
        log_activity(ActivityActions.PERSON_EDIT, request=request,
            description=f'افزودن سوال‌های اختصاصی برای «{person.full_name}»',
            target_type='survey', target_id=person.survey_id, target_repr=person.survey.title,
            metadata={'person_id': person.id, 'questions_count': len(questions_data), 'custom_questions': True})
        return Response(SurveyPersonSerializer(person, context={'request': request}).data)

class EmployeeSurveyListView(generics.ListAPIView):
    permission_classes = [IsEmployeeUser]
    serializer_class = SurveySerializer

    pagination_class = None

    def get_queryset(self):
        return (
            Survey.objects
            .filter(status__in=[Survey.STATUS_PUBLISHED, Survey.STATUS_CLOSED])
            .order_by('-published_at', '-created_at')
        )

    def list(self, request, *args, **kwargs):
        # Annotated counters + two grouped queries (bulk completions, this
        # user's answered rows) replace the former ~7 queries and a ratings
        # dump per survey.
        qs = (
            annotate_survey_list_stats(self.get_queryset())
            .select_related('created_by')
            .prefetch_related('people', 'questions')
        )
        surveys = list(qs)

        completed_counts = bulk_completed_response_counts([s.id for s in surveys])
        for survey in surveys:
            survey.bulk_total_responses = completed_counts.get(survey.id, 0)

        data = SurveySerializer(surveys, many=True, context={'request': request}).data

        answered_by_survey = defaultdict(dict)
        if surveys and any(s.active_questions_count > 0 for s in surveys):
            rows = (
                Rating.objects
                .filter(
                    survey_id__in=[s.id for s in surveys],
                    voter=request.user,
                    person__is_active=True,
                    question__is_active=True,
                )
                .values('survey_id', 'person_id')
                .annotate(answered_count=Count('question_id', distinct=True))
            )
            for row in rows:
                answered_by_survey[row['survey_id']][row['person_id']] = row['answered_count']

        for survey, item in zip(surveys, data):
            required_answers_per_person = survey.active_questions_count
            answered_by_person = answered_by_survey.get(survey.id, {})
            completed_person_ids = {
                person_id
                for person_id, answered_count in answered_by_person.items()
                if required_answers_per_person > 0 and answered_count == required_answers_per_person
            }

            item['my_votes_count'] = len(completed_person_ids)
            item['total_people'] = survey.active_people_count
            item['total_questions'] = survey.active_questions_count
        return Response(data)


class EmployeeSurveyDetailView(APIView):
    permission_classes = [IsEmployeeUser]

    def get(self, request, pk):
        try:
            survey = Survey.objects.get(pk=pk, status__in=[Survey.STATUS_PUBLISHED, Survey.STATUS_CLOSED])
        except Survey.DoesNotExist:
            return Response({'detail': 'نظرسنجی یافت نشد.'}, status=status.HTTP_404_NOT_FOUND)

        serializer = SurveyPublicSerializer(survey, context={'request': request})
        data = serializer.data

        completed_person_ids = participant_completed_person_ids(survey, voter_id=request.user.id)

        for person in data['people']:
            person['has_rated'] = person['id'] in completed_person_ids

        return Response(data)


class EmployeeRatePersonView(APIView):
    permission_classes = [IsEmployeeUser]

    def post(self, request, survey_id, person_id):
        try:
            survey = Survey.objects.get(pk=survey_id)
        except Survey.DoesNotExist:
            return Response({'detail': 'نظرسنجی یافت نشد.'}, status=status.HTTP_404_NOT_FOUND)

        if survey.status == Survey.STATUS_DRAFT:
            return Response({'detail': 'این نظرسنجی هنوز منتشر نشده است.'}, status=status.HTTP_400_BAD_REQUEST)

        if survey.status == Survey.STATUS_CLOSED:
            return Response({'detail': 'این نظرسنجی بسته شده است.'}, status=status.HTTP_400_BAD_REQUEST)

        try:
            person = SurveyPerson.objects.get(pk=person_id, survey=survey, is_active=True)
        except SurveyPerson.DoesNotExist:
            return Response({'detail': 'فرد مورد نظر یافت نشد.'}, status=status.HTTP_404_NOT_FOUND)

        questions = list(effective_questions_for_person(person))
        if not questions:
            return Response({'detail': 'این نظرسنجی هنوز سوال فعالی ندارد.'}, status=status.HTTP_400_BAD_REQUEST)

        existing_answer_count = Rating.objects.filter(
            survey=survey, person=person, voter=request.user,
            question__is_active=True,
        ).count()
        if existing_answer_count >= len(questions):
            return Response({'detail': 'شما قبلاً برای این فرد پاسخ ثبت کرده‌اید.'}, status=status.HTTP_400_BAD_REQUEST)

        serializer = RatingCreateSerializer(data=request.data)
        serializer.is_valid(raise_exception=True)

        submitted_answers = serializer.validated_data.get('answers')
        if submitted_answers is None:

            if len(questions) != 1:
                return Response({'detail': 'برای این نظرسنجی باید پاسخ همه سوال‌ها ارسال شود.'}, status=status.HTTP_400_BAD_REQUEST)
            submitted_answers = [{
                'question_id': questions[0].id,
                'score': serializer.validated_data.get('score'),
                'emoji_rating': serializer.validated_data.get('emoji_rating'),
                'comment': serializer.validated_data.get('comment'),
            }]

        try:
            validated_answers = validate_question_answers(questions, submitted_answers)
        except ValueError as exc:
            return Response({'detail': str(exc)}, status=status.HTTP_400_BAD_REQUEST)

        try:
            with transaction.atomic():
                Rating.objects.bulk_create([
                    Rating(
                        survey=survey,
                        person=person,
                        question=item['question'],
                        voter=request.user,
                        score=item['score'],
                        emoji_rating=item.get('emoji_rating'),
                        comment=item['comment'],
                        ip_address=get_client_ip(request),
                        user_agent=request.META.get('HTTP_USER_AGENT', '')[:500],
                    )
                    for item in validated_answers
                ])
        except IntegrityError:
            return Response({'detail': 'شما قبلاً برای این فرد پاسخ ثبت کرده‌اید.'}, status=status.HTTP_400_BAD_REQUEST)

        invalidate_survey_results(survey.id)
        invalidate_dashboard()
        invalidate_employee_survey_list(request.user.id)
        return Response({'detail': 'پاسخ‌های شما با موفقیت ثبت شد.'}, status=status.HTTP_201_CREATED)


class EmployeeMyRatingsView(APIView):
    permission_classes = [IsEmployeeUser]

    def get(self, request, survey_id):
        try:
            survey = Survey.objects.get(pk=survey_id)
        except Survey.DoesNotExist:
            return Response({'detail': 'نظرسنجی یافت نشد.'}, status=status.HTTP_404_NOT_FOUND)

        total_active_people = survey.people.filter(is_active=True).count()
        active_questions_count = survey.questions.filter(is_active=True).count()
        completed_person_ids = list(participant_completed_person_ids(survey, voter_id=request.user.id))

        return Response({
            'survey_id': survey.id,
            'rated_person_ids': completed_person_ids,
            'rated_count': len(completed_person_ids),
            'total_people': total_active_people,
            'total_questions': active_questions_count,
            'required_answers_count': len(required_question_pairs(survey)),
            'is_complete': len(completed_person_ids) == total_active_people and total_active_people > 0,
        })


class EmployeeSurveyResultsView(APIView):
    permission_classes = [IsEmployeeUser]

    def get(self, request, survey_id):
        try:
            survey = Survey.objects.get(pk=survey_id)
        except Survey.DoesNotExist:
            return Response({'detail': 'نظرسنجی یافت نشد.'}, status=status.HTTP_404_NOT_FOUND)

        if survey.results_visibility == Survey.VISIBILITY_ADMIN_ONLY:
            return Response({'detail': 'نتایج این نظرسنجی فقط برای مدیر قابل مشاهده است.'}, status=status.HTTP_403_FORBIDDEN)

        if survey.status != Survey.STATUS_CLOSED:
            return Response({'detail': 'نتایج پس از بسته شدن نظرسنجی در دسترس خواهد بود.'}, status=status.HTTP_403_FORBIDDEN)

        results = calculate_survey_results(survey, request)
        return Response({
            'survey': {
                'id': survey.id,
                'title': survey.title,
                'question': survey.question,
                'status': survey.status,
            },
            'results': results
        })



class AdminDashboardView(APIView):
    permission_classes = [IsAdminUser]

    def get(self, request):
        ck = key_dashboard()
        cached = cache.get(ck)
        if cached is not None:
            return Response(cached)

        from apps.accounts.models import User

        total_surveys = Survey.objects.count()
        draft_surveys = Survey.objects.filter(status=Survey.STATUS_DRAFT).count()
        published_surveys = Survey.objects.filter(status=Survey.STATUS_PUBLISHED).count()
        closed_surveys = Survey.objects.filter(status=Survey.STATUS_CLOSED).count()


        survey_meta = list(
            Survey.objects
            .annotate(
                ap=Count('people', filter=Q(people__is_active=True), distinct=True),
                aq=Count('questions', filter=Q(questions__is_active=True), distinct=True),
            )
            .values('id', 'ap', 'aq')
        )
        required_by_survey = {
            s['id']: s['ap'] * s['aq'] for s in survey_meta if s['ap'] * s['aq'] > 0
        }
        voter_answer_counts = (
            Rating.objects
            .filter(
                survey_id__in=required_by_survey.keys(),
                person__is_active=True,
                question__is_active=True,
            )
            .values('survey_id', 'voter_id')
            .annotate(answered_count=Count('id', distinct=True))
        )
        total_responses = sum(
            1
            for row in voter_answer_counts
            if row['answered_count'] >= required_by_survey.get(row['survey_id'], 0)
        )
        total_employees = User.objects.filter(role='employee').count()

        recent_surveys = list(
            annotate_survey_list_stats(Survey.objects.order_by('-created_at')).select_related('created_by')[:5]
        )
        recent_completed = bulk_completed_response_counts([s.id for s in recent_surveys])
        for survey in recent_surveys:
            survey.bulk_total_responses = recent_completed.get(survey.id, 0)
        recent_data = SurveySerializer(recent_surveys, many=True, context={'request': request}).data

        payload = {
            'stats': {
                'total_surveys': total_surveys,
                'draft_surveys': draft_surveys,
                'published_surveys': published_surveys,
                'closed_surveys': closed_surveys,
                'total_responses': total_responses,
                'total_employees': total_employees,
            },
            'recent_surveys': recent_data,
        }
        cache.set(ck, payload, settings.CACHE_TTL_DASHBOARD)
        return Response(payload)


class AdminDeleteAllDataView(APIView):
    """حذف تمام داده‌ها — فقط مدیر، غیرقابل بازگشت"""
    permission_classes = [IsAdminUser]

    def get(self, request):
        from apps.accounts.models import User

        return Response({
            'surveys': Survey.objects.count(),
            'people': SurveyPerson.objects.count(),
            'ratings': Rating.objects.count(),
            'employees': User.objects.filter(role='employee').count(),
        })

    def delete(self, request):
        from apps.accounts.models import User

        confirm = request.data.get('confirm')
        if confirm != 'DELETE_ALL':
            return Response(
                {'detail': 'برای تأیید، مقدار "DELETE_ALL" را ارسال کنید.'},
                status=status.HTTP_400_BAD_REQUEST
            )

        password = request.data.get('password')
        if not password or not request.user.check_password(password):
            return Response(
                {'detail': 'رمز عبور نادرست است.'},
                status=status.HTTP_400_BAD_REQUEST
            )

        ratings_count = Rating.objects.count()
        people_count = SurveyPerson.objects.count()
        surveys_count = Survey.objects.count()
        employees_count = User.objects.filter(role='employee').count()

        Rating.objects.all().delete()
        SurveyPerson.objects.all().delete()
        Survey.objects.all().delete()
        User.objects.filter(role='employee').delete()

        logger.warning(
            f"Admin {request.user.username} deleted ALL data: "
            f"{surveys_count} surveys, {people_count} survey-people, "
            f"{ratings_count} ratings, {employees_count} employees"
        )
        log_activity(
            ActivityActions.DELETE_ALL_DATA,
            request=request,
            description=(
                f'حذف تمام داده‌ها: {surveys_count} نظرسنجی، '
                f'{ratings_count} پاسخ، {employees_count} کارمند'
            ),
            target_type='system',
            target_repr='کل داده‌های سیستم',
            metadata={
                'surveys': surveys_count,
                'people': people_count,
                'ratings': ratings_count,
                'employees': employees_count,
            },
        )

        return Response({
            'detail': 'تمام داده‌ها با موفقیت حذف شدند.',
            'deleted': {
                'surveys': surveys_count,
                'people': people_count,
                'ratings': ratings_count,
                'employees': employees_count,
            }
        })


class AdminSurveyCommentsView(APIView):
    """Paginated comments for a specific person (and optionally question) in a survey.

    GET /admin/surveys/<pk>/comments/?person_id=<id>&question_id=<id>&page=1&page_size=20
    """
    permission_classes = [IsAdminUser]

    def get(self, request, pk):
        try:
            survey = Survey.objects.get(pk=pk)
        except Survey.DoesNotExist:
            return Response({'detail': 'نظرسنجی یافت نشد.'}, status=status.HTTP_404_NOT_FOUND)

        person_id   = request.query_params.get('person_id')
        question_id = request.query_params.get('question_id')
        try:
            page      = max(1, int(request.query_params.get('page', 1)))
            page_size = min(100, max(1, int(request.query_params.get('page_size', 20))))
        except (ValueError, TypeError):
            page, page_size = 1, 20

        # Completion must be judged within the same isolated section the
        # comments belong to (general group, or one particular person) -
        # never the whole-survey definition, or a particular person's
        # unfinished questions would hide otherwise-complete general comments.
        if person_id:
            person = get_object_or_404(SurveyPerson, pk=person_id, survey=survey)
            group_people = (
                [person] if not person.uses_default_questions
                else list(survey.people.filter(is_active=True, uses_default_questions=True))
            )
        elif question_id:
            question = get_object_or_404(SurveyQuestion, pk=question_id, survey=survey)
            group_people = (
                [question.person] if question.person_id
                else list(survey.people.filter(is_active=True, uses_default_questions=True))
            )
        else:
            group_people = list(survey.people.filter(is_active=True))

        completed_voter_ids, completed_anon_tokens = completed_participants_for(survey, group_people)

        qs = Rating.objects.filter(
            survey=survey,
            person__is_active=True,
            question__is_active=True,
        ).filter(
            Q(voter_id__in=completed_voter_ids) | Q(anonymous_token__in=completed_anon_tokens)
        ).exclude(comment__isnull=True).exclude(comment__exact='')

        if person_id:
            qs = qs.filter(person_id=person_id)
        if question_id:
            qs = qs.filter(question_id=question_id)

        qs = qs.order_by('person_id', 'question__display_order', 'created_at')

        total = qs.count()
        offset = (page - 1) * page_size
        comments_page = list(qs.values('comment', 'question__text')[offset:offset + page_size])

        return Response({
            'total': total,
            'page': page,
            'page_size': page_size,
            'total_pages': (total + page_size - 1) // page_size,
            'comments': [
                {'comment': row['comment'], 'question_text': row['question__text']}
                for row in comments_page
            ],
        })



class AdminHashLinkListCreateView(APIView):
    """List all hash links for a survey, or create a new one."""
    permission_classes = [IsAdminUser]

    def get(self, request, survey_id):
        try:
            survey = Survey.objects.get(pk=survey_id)
        except Survey.DoesNotExist:
            return Response({'detail': 'نظرسنجی یافت نشد.'}, status=status.HTTP_404_NOT_FOUND)

        cache_key = key_hash_links(survey.id)
        cached = cache.get(cache_key)
        if cached is not None:
            return Response(cached)

        links = survey.hash_links.all().order_by('-created_at')
        data = SurveyHashLinkSerializer(links, many=True).data
        try:
            cache.set(cache_key, data, timeout=60)
        except Exception:
            logger.debug('cache.set for hash_links(%s) failed silently', survey.id)
        return Response(data)

    def post(self, request, survey_id):
        try:
            survey = Survey.objects.get(pk=survey_id)
        except Survey.DoesNotExist:
            return Response({'detail': 'نظرسنجی یافت نشد.'}, status=status.HTTP_404_NOT_FOUND)

        label = (request.data.get('label') or '').strip()
        if label and survey.hash_links.filter(label__iexact=label).exists():
            return Response(
                {'detail': 'برای این نظرسنجی لینکی با همین نام وجود دارد.'},
                status=status.HTTP_400_BAD_REQUEST,
            )

        serializer = SurveyHashLinkSerializer(data={
            'label': label,
            'max_participants': request.data.get('max_participants'),
            'expiry_value': request.data.get('expiry_value'),
            'expiry_unit': request.data.get('expiry_unit'),
        })
        serializer.is_valid(raise_exception=True)
        link = SurveyHashLink.objects.create(
            survey=survey,
            label=label,
            max_participants=serializer.validated_data.get('max_participants'),
            expiry_value=serializer.validated_data.get('expiry_value'),
            expiry_unit=serializer.validated_data.get('expiry_unit'),
        )

        log_activity(
            ActivityActions.HASH_LINK_CREATE,
            request=request,
            description=f'ایجاد لینک هش برای نظرسنجی «{survey.title}»',
            target_type='survey',
            target_id=survey.id,
            target_repr=survey.title,
            metadata={
                'token': link.token, 'label': label,
                'max_participants': link.max_participants,
                'expiry_value': link.expiry_value,
                'expiry_unit': link.expiry_unit,
            },
        )
        invalidate_dashboard()
        invalidate_hash_links(survey.id)
        return Response(SurveyHashLinkSerializer(link).data, status=status.HTTP_201_CREATED)


class AdminHashLinkDetailView(APIView):
    """Update (label/active) or delete a specific hash link."""
    permission_classes = [IsAdminUser]

    def patch(self, request, pk):
        try:
            link = SurveyHashLink.objects.select_related('survey').get(pk=pk)
        except SurveyHashLink.DoesNotExist:
            return Response({'detail': 'لینک هش یافت نشد.'}, status=status.HTTP_404_NOT_FOUND)

        old_active = link.is_active
        limits_changed = False

        if 'label' in request.data:
            link.label = (request.data['label'] or '').strip()
        if 'is_active' in request.data:
            link.is_active = bool(request.data['is_active'])

        limit_fields = {}
        for field in ('max_participants', 'expiry_value', 'expiry_unit'):
            if field in request.data:
                limit_fields[field] = request.data[field]
                limits_changed = True

        if limit_fields:
            serializer = SurveyHashLinkSerializer(link, data=limit_fields, partial=True)
            serializer.is_valid(raise_exception=True)
            for field, value in limit_fields.items():
                setattr(link, field, serializer.validated_data.get(field) if field in serializer.validated_data else value)

        link.save()

        if old_active != link.is_active:
            state = 'فعال' if link.is_active else 'غیرفعال'
            log_activity(
                ActivityActions.HASH_LINK_TOGGLE,
                request=request,
                description=f'لینک هش نظرسنجی «{link.survey.title}» {state} شد',
                target_type='survey',
                target_id=link.survey.id,
                target_repr=link.survey.title,
                metadata={'token': link.token, 'is_active': link.is_active},
            )

        if limits_changed:
            log_activity(
                ActivityActions.HASH_LINK_UPDATE_LIMITS,
                request=request,
                description=f'محدودیت‌های لینک هش نظرسنجی «{link.survey.title}» تغییر کرد',
                target_type='survey',
                target_id=link.survey.id,
                target_repr=link.survey.title,
                metadata={
                    'token': link.token,
                    'max_participants': link.max_participants,
                    'expiry_value': link.expiry_value,
                    'expiry_unit': link.expiry_unit,
                },
            )

        invalidate_hash_links(link.survey.id)
        return Response(SurveyHashLinkSerializer(link).data)

    def delete(self, request, pk):
        try:
            link = SurveyHashLink.objects.select_related('survey').get(pk=pk)
        except SurveyHashLink.DoesNotExist:
            return Response({'detail': 'لینک هش یافت نشد.'}, status=status.HTTP_404_NOT_FOUND)

        survey = link.survey
        token = link.token
        link.delete()

        log_activity(
            ActivityActions.HASH_LINK_DELETE,
            request=request,
            description=f'حذف لینک هش نظرسنجی «{survey.title}»',
            target_type='survey',
            target_id=survey.id,
            target_repr=survey.title,
            metadata={'token': token},
        )
        invalidate_dashboard()
        invalidate_hash_links(survey.id)
        return Response(status=status.HTTP_204_NO_CONTENT)


class AnonymousSurveyDetailView(APIView):
    """Return survey detail for an anonymous participant using a hash token.
    Also returns ip_locked=True when this device may no longer participate:
    either it already finished the survey, or its registration is bound to a
    different anonymous token than the one presented. An unfinished session
    presenting its own token stays unlockable so participants can resume.
    """
    permission_classes = [AllowAny]
    throttle_classes = [AnonymousSurveyRateThrottle]

    def get(self, request, token):
        try:
            link = SurveyHashLink.objects.select_related('survey').get(token=token)
        except SurveyHashLink.DoesNotExist:
            return Response({'detail': 'لینک معتبر نیست.'}, status=status.HTTP_404_NOT_FOUND)

        if not link.is_active:
            return Response({'detail': 'این لینک غیرفعال شده است.'}, status=status.HTTP_403_FORBIDDEN)

        if link.is_expired:
            return Response({'detail': 'مهلت استفاده از این لینک به پایان رسیده است.'}, status=status.HTTP_403_FORBIDDEN)

        survey = link.survey
        if survey.status == Survey.STATUS_DRAFT:
            return Response({'detail': 'این نظرسنجی هنوز منتشر نشده است.'}, status=status.HTTP_400_BAD_REQUEST)
        if survey.status == Survey.STATUS_CLOSED:
            return Response({'detail': 'این نظرسنجی بسته شده است.'}, status=status.HTTP_400_BAD_REQUEST)

        client_ip = get_client_ip(request)
        presented_token = request.query_params.get('anonymous_token', '').strip()
        ip_locked = False
        if client_ip:
            registration = AnonymousParticipation.objects.filter(
                survey=survey, hash_link=link, ip_address=client_ip
            ).first()
            if registration is not None:
                if registration.finished_at:
                    ip_locked = True
                else:
                    # Unfinished device: only the bound token may continue.
                    ip_locked = not presented_token or presented_token != registration.anonymous_token

        if link.is_full and not ip_locked:
            return Response({'detail': 'ظرفیت شرکت‌کنندگان این لینک تکمیل شده است.'}, status=status.HTTP_403_FORBIDDEN)

        from .serializers import SurveyPublicSerializer
        data = SurveyPublicSerializer(survey, context={'request': request}).data
        data['ip_locked'] = ip_locked
        return Response(data)


class AnonymousRatePersonView(APIView):
    """Anonymous participation endpoint — no authentication required."""
    permission_classes = [AllowAny]
    throttle_classes = [AnonymousSurveyRateThrottle]

    def post(self, request, token, person_id):
        try:
            link = SurveyHashLink.objects.select_related('survey').get(token=token)
        except SurveyHashLink.DoesNotExist:
            return Response({'detail': 'لینک معتبر نیست.'}, status=status.HTTP_404_NOT_FOUND)

        if not link.is_active:
            return Response({'detail': 'این لینک غیرفعال شده است.'}, status=status.HTTP_403_FORBIDDEN)

        if link.is_expired:
            return Response({'detail': 'مهلت استفاده از این لینک به پایان رسیده است.'}, status=status.HTTP_403_FORBIDDEN)

        survey = link.survey
        client_ip = get_client_ip(request)
        if survey.status == Survey.STATUS_DRAFT:
            return Response({'detail': 'این نظرسنجی هنوز منتشر نشده است.'}, status=status.HTTP_400_BAD_REQUEST)
        if survey.status == Survey.STATUS_CLOSED:
            return Response({'detail': 'این نظرسنجی بسته شده است.'}, status=status.HTTP_400_BAD_REQUEST)
        anon_session = request.data.get('anonymous_token', '').strip()
        if not anon_session or len(anon_session) > 64:
            return Response({'detail': 'توکن ناشناس معتبر نیست.'}, status=status.HTTP_400_BAD_REQUEST)

        # Device binding: the first ballot from this IP registers a
        # participation bound to anon_session. Any later ballot from the same
        # IP must present that same token, so one device can never split
        # ballots across several tokens (the partial-ballot loophole).
        registration = None
        resuming_registration = False
        if client_ip:
            registration = AnonymousParticipation.objects.filter(
                survey=survey,
                hash_link=link,
                ip_address=client_ip,
            ).first()
            if registration is not None:
                if registration.anonymous_token != anon_session:
                    return Response({'detail': 'از این دستگاه/آدرس IP قبلا در این نظرسنجی شرکت شده است.'}, status=status.HTTP_400_BAD_REQUEST)
                resuming_registration = True

        if link.is_full and not resuming_registration:
            # An in-progress session keeps its already-consumed slot and may
            # finish; only brand-new participants are capped by the limit.
            return Response({'detail': 'ظرفیت شرکت‌کنندگان این لینک تکمیل شده است.'}, status=status.HTTP_403_FORBIDDEN)
        try:
            person = SurveyPerson.objects.get(pk=person_id, survey=survey, is_active=True)
        except SurveyPerson.DoesNotExist:
            return Response({'detail': 'فرد مورد نظر یافت نشد.'}, status=status.HTTP_404_NOT_FOUND)

        questions = list(effective_questions_for_person(person))
        if not questions:
            return Response({'detail': 'این نظرسنجی هنوز سوال فعالی ندارد.'}, status=status.HTTP_400_BAD_REQUEST)
        existing_answer_count = Rating.objects.filter(
            survey=survey, person=person,
            anonymous_token=anon_session,
            question__is_active=True,
        ).count()
        if existing_answer_count >= len(questions):
            return Response({'detail': 'شما قبلاً برای این فرد پاسخ ثبت کرده‌اید.'}, status=status.HTTP_400_BAD_REQUEST)

        from .serializers import RatingCreateSerializer
        serializer = RatingCreateSerializer(data=request.data)
        serializer.is_valid(raise_exception=True)

        submitted_answers = serializer.validated_data.get('answers')
        if submitted_answers is None:
            if len(questions) != 1:
                return Response({'detail': 'برای این نظرسنجی باید پاسخ همه سوال‌ها ارسال شود.'}, status=status.HTTP_400_BAD_REQUEST)
            submitted_answers = [{
                'question_id': questions[0].id,
                'score': serializer.validated_data.get('score'),
                'emoji_rating': serializer.validated_data.get('emoji_rating'),
                'comment': serializer.validated_data.get('comment'),
            }]

        try:
            validated_answers = validate_question_answers(questions, submitted_answers)
        except ValueError as exc:
            return Response({'detail': str(exc)}, status=status.HTTP_400_BAD_REQUEST)

        participation_finished = False
        try:
            with transaction.atomic():
                # Serialize submissions for this link so capacity checks and
                # device registrations cannot race, and roll answers back if
                # the binding check below rejects a concurrent new session.
                locked_link = SurveyHashLink.objects.select_for_update().get(pk=link.pk)

                current_registration = None
                if client_ip:
                    # Re-check under lock: two fresh tokens racing from one
                    # IP must not both register.
                    current_registration = AnonymousParticipation.objects.filter(
                        survey=survey, hash_link=locked_link, ip_address=client_ip,
                    ).first()
                    if (
                        current_registration is not None
                        and current_registration.anonymous_token != anon_session
                    ):
                        return Response({'detail': 'از این دستگاه/آدرس IP قبلا در این نظرسنجی شرکت شده است.'}, status=status.HTTP_400_BAD_REQUEST)
                if locked_link.is_full and not resuming_registration:
                    return Response({'detail': 'ظرفیت شرکت‌کنندگان این لینک تکمیل شده است.'}, status=status.HTTP_403_FORBIDDEN)

                Rating.objects.bulk_create([
                    Rating(
                        survey=survey,
                        person=person,
                        question=item['question'],
                        voter=None,  # anonymous — no user
                        anonymous_token=anon_session,
                        score=item['score'],
                        emoji_rating=item.get('emoji_rating'),
                        comment=item['comment'],
                        ip_address=client_ip,
                        user_agent=request.META.get('HTTP_USER_AGENT', '')[:500],
                    )
                    for item in validated_answers
                ])

                required = len(required_question_pairs(survey))
                answered_count = Rating.objects.filter(
                    survey=survey,
                    anonymous_token=anon_session,
                    person__is_active=True,
                    question__is_active=True,
                ).count()

                if client_ip:
                    if current_registration is None:
                        AnonymousParticipation.objects.create(
                            survey=survey,
                            hash_link=locked_link,
                            ip_address=client_ip,
                            anonymous_token=anon_session,
                            user_agent=request.META.get('HTTP_USER_AGENT', '')[:500],
                        )
                        SurveyHashLink.objects.filter(pk=locked_link.pk).update(
                            anonymous_participant_count=F('anonymous_participant_count') + 1
                        )
                    if answered_count >= required:
                        # Mark completion exactly once, even across resumes.
                        participation_finished = bool(AnonymousParticipation.objects.filter(
                            survey=survey,
                            hash_link=locked_link,
                            ip_address=client_ip,
                            finished_at__isnull=True,
                        ).update(finished_at=timezone.now()))
        except IntegrityError:
            return Response({'detail': 'شما قبلاً برای این فرد پاسخ ثبت کرده‌اید.'}, status=status.HTTP_400_BAD_REQUEST)
        if participation_finished:
            log_activity(
                ActivityActions.ANONYMOUS_VOTE,
                request=request,
                description=f'یک شرکت‌کننده ناشناس نظرسنجی «{survey.title}» را تکمیل کرد',
                target_type='survey',
                target_id=survey.id,
                target_repr=survey.title,
                metadata={'token': token, 'anonymous_ip': client_ip},
            )

        invalidate_survey_results(survey.id)
        invalidate_dashboard()
        return Response({'detail': 'پاسخ‌های شما با موفقیت ثبت شد.'}, status=status.HTTP_201_CREATED)


class AnonymousMyRatingsView(APIView):
    """Return which people an anonymous session has already rated."""
    permission_classes = [AllowAny]
    throttle_classes = [AnonymousSurveyRateThrottle]

    def get(self, request, token, survey_id):
        try:
            link = SurveyHashLink.objects.select_related('survey').get(token=token)
        except SurveyHashLink.DoesNotExist:
            return Response({'detail': 'لینک معتبر نیست.'}, status=status.HTTP_404_NOT_FOUND)

        if not link.is_active:
            return Response({'detail': 'این لینک غیرفعال شده است.'}, status=status.HTTP_403_FORBIDDEN)

        survey = link.survey
        if survey.id != survey_id:
            return Response({'detail': 'نظرسنجی یافت نشد.'}, status=status.HTTP_404_NOT_FOUND)

        total_active_people = survey.people.filter(is_active=True).count()
        active_questions_count = survey.questions.filter(is_active=True).count()
        anon_session = request.query_params.get('anonymous_token', '').strip()

        client_ip = get_client_ip(request)
        registration = None
        if client_ip:
            registration = AnonymousParticipation.objects.filter(
                survey=survey,
                hash_link=link,
                ip_address=client_ip,
            ).first()
        if registration is not None:
            # Locked when finished, or when this device is registered under a
            # different token than the one presented. A same-token unfinished
            # session falls through so real progress keeps being reported.
            if registration.finished_at or anon_session != registration.anonymous_token:
                return Response({
                    'survey_id': survey.id,
                    'rated_person_ids': list(survey.people.filter(is_active=True).values_list('id', flat=True)),
                    'rated_count': total_active_people,
                    'total_people': total_active_people,
                    'total_questions': active_questions_count,
                    'required_answers_count': len(required_question_pairs(survey)),
                    'is_complete': total_active_people > 0,
                    'ip_locked': True,
                })

        if not anon_session:
            return Response({'rated_person_ids': [], 'rated_count': 0, 'total_people': 0, 'is_complete': False, 'ip_locked': False})

        completed_person_ids = list(participant_completed_person_ids(survey, anonymous_token=anon_session))

        return Response({
            'survey_id': survey.id,
            'rated_person_ids': completed_person_ids,
            'rated_count': len(completed_person_ids),
            'total_people': total_active_people,
            'total_questions': active_questions_count,
            'required_answers_count': len(required_question_pairs(survey)),
            'is_complete': len(completed_person_ids) == total_active_people and total_active_people > 0,
            'ip_locked': False,
        })
