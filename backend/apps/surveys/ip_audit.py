"""Query and Excel helpers for the admin IP Response Audit feature."""

import io
from collections import OrderedDict

from django.core.paginator import Paginator
from django.db.models import Case, CharField, Count, F, Max, Q, Value, When
from django.db.models.functions import Cast, Concat
from django.utils import timezone

from apps.core.export_security import sanitize_cell

from .models import Rating, SurveyPerson


EMOJI_VISUALS = {
    Rating.EMOJI_BAD: '😞',
    Rating.EMOJI_AVERAGE: '😐',
    Rating.EMOJI_GOOD: '🙂',
    Rating.EMOJI_EXCELLENT: '😍',
}


def _submission_identifier(rating):
    if rating.voter_id:
        return f'user:{rating.voter_id}'
    if rating.anonymous_token:
        return f'anonymous:{rating.anonymous_token}'
    return f'legacy-rating:{rating.id}'


def _question_type(question):
    kinds = []
    if question.has_score:
        kinds.append('numeric')
    if question.has_emoji:
        kinds.append('emoji')
    if question.has_comment:
        kinds.append('text')
    return '+'.join(kinds) or 'unknown'


def _submission_key_expression():
    return Case(
        When(
            voter_id__isnull=False,
            then=Concat(Value('user:'), Cast('voter_id', CharField())),
        ),
        When(
            Q(anonymous_token__isnull=False) & ~Q(anonymous_token=''),
            then=Concat(Value('anonymous:'), F('anonymous_token')),
        ),
        default=Concat(Value('legacy-rating:'), Cast('id', CharField())),
        output_field=CharField(),
    )


def audit_ratings_queryset(survey, ip_address):
    return (
        Rating.objects
        .filter(survey=survey, ip_address=ip_address)
        .select_related('person', 'question', 'voter')
        .order_by(
            'person__display_order', 'person__id',
            'created_at', 'question__display_order', 'question__id', 'id',
        )
    )


def _audit_summary(survey, ip_address):
    return (
        Rating.objects
        .filter(survey=survey, ip_address=ip_address)
        .aggregate(
            total_answers=Count('id'),
            total_linked_submissions=Count(
                _submission_key_expression(), distinct=True,
            ),
            total_surveyed_people=Count('person_id', distinct=True),
            latest_submission_at=Max('created_at'),
        )
    )


def build_ip_audit_payload(survey, ip_address, page=None, page_size=None):
    """Return ratings grouped without ever merging different evaluated people."""
    pagination = None
    ratings_queryset = audit_ratings_queryset(survey, ip_address)
    if page is not None and page_size is not None:
        people_queryset = (
            SurveyPerson.objects
            .filter(
                survey=survey,
                ratings__survey=survey,
                ratings__ip_address=ip_address,
            )
            .distinct()
            .order_by('display_order', 'id')
        )
        people_page = Paginator(people_queryset, page_size).get_page(page)
        person_ids = [person.id for person in people_page.object_list]
        ratings_queryset = ratings_queryset.filter(person_id__in=person_ids)
        pagination = {
            'page': people_page.number,
            'page_size': page_size,
            'total': people_page.paginator.count,
            'total_pages': people_page.paginator.num_pages,
            'has_previous': people_page.has_previous(),
            'has_next': people_page.has_next(),
        }

    ratings = list(ratings_queryset)
    people = OrderedDict()

    for rating in ratings:
        identifier = _submission_identifier(rating)

        person_group = people.setdefault(rating.person_id, {
            'surveyed_person_id': rating.person_id,
            'surveyed_person_name': rating.person.full_name,
            'role_title': rating.person.role_title,
            'department': rating.person.department,
            'submissions': OrderedDict(),
        })
        submission = person_group['submissions'].setdefault(identifier, {
            'submission_identifier': identifier,
            'submitted_at': rating.created_at,
            'answers': [],
        })
        if rating.created_at > submission['submitted_at']:
            submission['submitted_at'] = rating.created_at

        submission['answers'].append({
            'selected_ip_address': str(ip_address),
            'survey_id': survey.id,
            'survey_title': survey.title,
            'submission_identifier': identifier,
            'submitted_at': rating.created_at,
            'surveyed_person_id': rating.person_id,
            'surveyed_person_name': rating.person.full_name,
            'question_id': rating.question_id,
            'question_order': rating.question.display_order,
            'question_text': rating.question.text,
            'question_type': _question_type(rating.question),
            'numeric_score': rating.score,
            'emoji_rating': rating.emoji_rating,
            'emoji_label': rating.get_emoji_rating_display() if rating.emoji_rating else None,
            'free_text_answer': rating.comment or None,
        })

    serialized_people = []
    for person in people.values():
        person['submissions'] = list(person['submissions'].values())
        serialized_people.append(person)

    payload = {
        'survey': {'id': survey.id, 'title': survey.title},
        'selected_ip_address': str(ip_address),
        'summary': _audit_summary(survey, ip_address),
        'people': serialized_people,
    }
    if pagination is not None:
        payload['pagination'] = pagination
    return payload


def available_ip_payload(survey, search='', page=1, page_size=8):
    """Return a paginated, searchable list of IPs represented in this survey."""
    ratings = Rating.objects.filter(survey=survey, ip_address__isnull=False)
    if search:
        ratings = ratings.filter(ip_address__icontains=search)

    grouped = (
        ratings
        .values('ip_address')
        .annotate(
            response_count=Count('id'),
            surveyed_person_count=Count('person_id', distinct=True),
            submission_count=Count(_submission_key_expression(), distinct=True),
            latest_submission_at=Max('created_at'),
        )
        .order_by('-latest_submission_at', 'ip_address')
    )
    ip_page = Paginator(grouped, page_size).get_page(page)
    items = [
        {
            'ip_address': str(row['ip_address']),
            'response_count': row['response_count'],
            'surveyed_person_count': row['surveyed_person_count'],
            'submission_count': row['submission_count'],
            'latest_submission_at': row['latest_submission_at'],
        }
        for row in ip_page.object_list
    ]
    return {
        'ips': items,
        'pagination': {
            'page': ip_page.number,
            'page_size': page_size,
            'total': ip_page.paginator.count,
            'total_pages': ip_page.paginator.num_pages,
            'has_previous': ip_page.has_previous(),
            'has_next': ip_page.has_next(),
        },
    }


def build_ip_audit_workbook(survey, ip_address, payload):
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    from openpyxl.utils import get_column_letter

    workbook = Workbook()
    workbook.properties.creator = 'InsightFlow'
    workbook.properties.title = f'IP Response Audit - {survey.title}'
    workbook.properties.subject = f'Responses submitted from {ip_address}'
    sheet = workbook.active
    sheet.title = 'ممیزی پاسخ IP'
    sheet.sheet_view.rightToLeft = True
    sheet.freeze_panes = 'A7'
    sheet.sheet_view.showGridLines = False
    sheet.sheet_properties.tabColor = '4F46E5'
    sheet.sheet_properties.pageSetUpPr.fitToPage = True
    sheet.page_setup.orientation = 'landscape'
    sheet.page_setup.fitToWidth = 1
    sheet.page_setup.fitToHeight = 0
    sheet.print_title_rows = '6:6'
    sheet.oddFooter.center.text = 'InsightFlow • IP Response Audit'
    sheet.oddFooter.right.text = 'Page &P of &N'

    title_fill = PatternFill('solid', fgColor='4F46E5')
    header_fill = PatternFill('solid', fgColor='1E293B')
    metadata_label_fill = PatternFill('solid', fgColor='E0E7FF')
    metadata_value_fill = PatternFill('solid', fgColor='F8FAFC')
    summary_fill = PatternFill('solid', fgColor='EEF2FF')
    alternate_fill = PatternFill('solid', fgColor='F8FAFC')
    score_low_fill = PatternFill('solid', fgColor='FEF2F2')
    score_mid_fill = PatternFill('solid', fgColor='FFFBEB')
    score_high_fill = PatternFill('solid', fgColor='F0FDF4')
    emoji_fills = {
        Rating.EMOJI_BAD: score_low_fill,
        Rating.EMOJI_AVERAGE: score_mid_fill,
        Rating.EMOJI_GOOD: PatternFill('solid', fgColor='F7FEE7'),
        Rating.EMOJI_EXCELLENT: score_high_fill,
    }
    thin = Side(style='thin', color='E2E8F0')
    group_side = Side(style='medium', color='A5B4FC')
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    right = Alignment(horizontal='right', vertical='center', wrap_text=True)
    center = Alignment(horizontal='center', vertical='center', wrap_text=True)
    top_right = Alignment(horizontal='right', vertical='top', wrap_text=True)

    sheet.append([sanitize_cell(f'ممیزی پاسخ‌های IP — {survey.title}')])
    sheet.merge_cells('A1:J1')
    for cell in sheet[1]:
        cell.fill = title_fill
    sheet['A1'].font = Font(bold=True, color='FFFFFF', size=14)
    sheet['A1'].alignment = center
    sheet.row_dimensions[1].height = 34

    sheet.append(['عنوان نظرسنجی', sanitize_cell(survey.title)])
    sheet.append(['آدرس IP انتخاب‌شده', sanitize_cell(str(ip_address))])
    sheet.append(['زمان تولید خروجی', timezone.localtime().strftime('%Y-%m-%d %H:%M:%S')])
    for row_number in range(2, 5):
        sheet.merge_cells(
            start_row=row_number, start_column=2,
            end_row=row_number, end_column=10,
        )
        label = sheet.cell(row=row_number, column=1)
        label.fill = metadata_label_fill
        label.font = Font(bold=True, color='3730A3')
        label.alignment = right
        label.border = border
        value = sheet.cell(row=row_number, column=2)
        value.fill = metadata_value_fill
        value.font = Font(color='334155')
        value.alignment = right
        value.border = border
        for column in range(3, 11):
            sheet.cell(row=row_number, column=column).fill = metadata_value_fill
            sheet.cell(row=row_number, column=column).border = border
        sheet.row_dimensions[row_number].height = 22

    summary = payload['summary']
    summary_text = (
        f"{summary['total_answers']} پاسخ  •  "
        f"{summary['total_linked_submissions']} مشارکت  •  "
        f"{summary['total_surveyed_people']} فرد ارزیابی‌شده"
    )
    sheet.append([summary_text])
    sheet.merge_cells('A5:J5')
    for cell in sheet[5]:
        cell.fill = summary_fill
    sheet['A5'].font = Font(bold=True, color='4338CA', size=10)
    sheet['A5'].alignment = center
    sheet.row_dimensions[5].height = 24

    headers = [
        'IP address', 'submission identifier', 'submitted at', 'surveyed person',
        'question order', 'question text', 'answer type', 'numeric score',
        'emoji rating', 'free-text answer',
    ]
    sheet.append(headers)
    for cell in sheet[6]:
        cell.fill = header_fill
        cell.font = Font(bold=True, color='FFFFFF')
        cell.alignment = center
        cell.border = border
    sheet.row_dimensions[6].height = 30

    previous_person_id = None
    data_row_number = 0
    for person in payload['people']:
        for submission in person['submissions']:
            for answer in submission['answers']:
                data_row_number += 1
                emoji = ''
                if answer['emoji_rating']:
                    emoji = (
                        f"{EMOJI_VISUALS.get(answer['emoji_rating'], '')} "
                        f"{answer['emoji_label'] or answer['emoji_rating']}"
                    ).strip()
                sheet.append([
                    sanitize_cell(answer['selected_ip_address']),
                    sanitize_cell(answer['submission_identifier']),
                    timezone.localtime(answer['submitted_at']).strftime('%Y-%m-%d %H:%M:%S'),
                    sanitize_cell(answer['surveyed_person_name']),
                    answer['question_order'],
                    sanitize_cell(answer['question_text']),
                    answer['question_type'],
                    answer['numeric_score'] if answer['numeric_score'] is not None else '',
                    sanitize_cell(emoji),
                    sanitize_cell(answer['free_text_answer'] or ''),
                ])
                row_number = sheet.max_row
                is_new_person = answer['surveyed_person_id'] != previous_person_id
                row_border = Border(
                    left=thin,
                    right=thin,
                    top=group_side if is_new_person else thin,
                    bottom=thin,
                )
                for column, cell in enumerate(sheet[row_number], 1):
                    cell.border = row_border
                    cell.alignment = (
                        center if column in (1, 3, 5, 7, 8, 9) else top_right
                    )
                    if data_row_number % 2 == 0:
                        cell.fill = alternate_fill

                score_cell = sheet.cell(row=row_number, column=8)
                score = answer['numeric_score']
                if score is not None:
                    if score < 4:
                        score_cell.fill = score_low_fill
                        score_cell.font = Font(bold=True, color='DC2626')
                    elif score < 7:
                        score_cell.fill = score_mid_fill
                        score_cell.font = Font(bold=True, color='D97706')
                    else:
                        score_cell.fill = score_high_fill
                        score_cell.font = Font(bold=True, color='059669')

                emoji_cell = sheet.cell(row=row_number, column=9)
                emoji_key = answer['emoji_rating']
                if emoji_key:
                    emoji_cell.fill = emoji_fills.get(emoji_key, metadata_value_fill)
                    emoji_cell.font = Font(bold=True, color='334155')

                comment_length = len(answer['free_text_answer'] or '')
                sheet.row_dimensions[row_number].height = min(
                    72, max(22, 15 + (comment_length // 55) * 12),
                )
                previous_person_id = answer['surveyed_person_id']

    widths = [20, 28, 21, 26, 14, 45, 18, 14, 20, 55]
    for index, width in enumerate(widths, 1):
        sheet.column_dimensions[get_column_letter(index)].width = width
    sheet.auto_filter.ref = f'A6:J{max(sheet.max_row, 6)}'
    sheet.print_area = f'A1:J{max(sheet.max_row, 6)}'
    sheet.sheet_properties.outlinePr.summaryBelow = False
    sheet.page_margins.left = 0.25
    sheet.page_margins.right = 0.25
    sheet.page_margins.top = 0.5
    sheet.page_margins.bottom = 0.5

    output = io.BytesIO()
    workbook.save(output)
    return output.getvalue()
