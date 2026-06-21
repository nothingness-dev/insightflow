"""Export center for activity logs (CSV / Excel / PDF).

All three formats consume the SAME filtered queryset and a date range. To stay
scalable the queryset is streamed/iterated lazily and the PDF is row-capped
(PDF is an executive summary; CSV/Excel carry full fidelity).
"""
import csv
import io

from django.utils import timezone

from .models import ActivityLog


# PDF stays readable & bounded; CSV/Excel keep everything.
PDF_MAX_ROWS = 1500
EXCEL_CELL_LIMIT = 32_000

_STATUS_LABELS = {
    ActivityLog.STATUS_SUCCESS: 'موفق',
    ActivityLog.STATUS_FAILED: 'ناموفق',
}

# Order of columns shared across formats.
_HEADERS = [
    'زمان', 'نوع فعالیت', 'کاربر', 'نام کاربری', 'نقش',
    'وضعیت', 'حساس', 'شرح', 'نوع هدف', 'عنوان هدف', 'آدرس IP',
]


def _fmt_dt(value):
    if not value:
        return ''
    return timezone.localtime(value).strftime('%Y-%m-%d %H:%M:%S')


def _role_label(role):
    return {'admin': 'مدیر', 'employee': 'کارمند'}.get(role, role or '')


def _row_values(log):
    return [
        _fmt_dt(log.created_at),
        log.action_label,
        log.actor_full_name or '—',
        log.actor_username or '—',
        _role_label(log.actor_role),
        _STATUS_LABELS.get(log.status, log.status),
        'بله' if log.is_critical else 'خیر',
        log.description or '',
        log.target_type or '',
        log.target_repr or '',
        log.ip_address or '',
    ]


def _filename(export_format, date_from, date_to):
    ext = {'csv': 'csv', 'excel': 'xlsx', 'pdf': 'pdf'}[export_format]
    return f'activity_logs_{date_from:%Y%m%d}_{date_to:%Y%m%d}.{ext}'


# ───────────────────────── CSV ─────────────────────────
def _build_csv(qs):
    buffer = io.StringIO()
    writer = csv.writer(buffer)
    writer.writerow(_HEADERS)
    for log in qs.iterator(chunk_size=500):
        writer.writerow(_row_values(log))
    # BOM so Excel opens the UTF-8 file with correct Persian glyphs.
    return ('\ufeff' + buffer.getvalue()).encode('utf-8')


# ───────────────────────── Excel ─────────────────────────
def _build_excel(qs):
    try:
        import openpyxl
        from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
        from openpyxl.utils import get_column_letter
    except ImportError as exc:  # pragma: no cover
        raise RuntimeError('خروجی Excel در دسترس نیست.') from exc

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'گزارش فعالیت‌ها'
    ws.sheet_view.rightToLeft = True

    header_fill = PatternFill('solid', fgColor='1E293B')
    header_font = Font(bold=True, color='FFFFFF', size=10)
    critical_fill = PatternFill('solid', fgColor='FEF2F2')
    failed_font = Font(color='B91C1C', bold=True)
    center = Alignment(horizontal='center', vertical='center', wrap_text=True)
    thin = Side(style='thin', color='E2E8F0')
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    ws.append(_HEADERS)
    for col_idx, _ in enumerate(_HEADERS, 1):
        cell = ws.cell(row=1, column=col_idx)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = center
        cell.border = border

    for log in qs.iterator(chunk_size=500):
        values = [str(v)[:EXCEL_CELL_LIMIT] for v in _row_values(log)]
        ws.append(values)
        row_idx = ws.max_row
        for col_idx in range(1, len(_HEADERS) + 1):
            cell = ws.cell(row=row_idx, column=col_idx)
            cell.alignment = center
            cell.border = border
            if log.is_critical:
                cell.fill = critical_fill
            if log.status == ActivityLog.STATUS_FAILED and col_idx == 6:
                cell.font = failed_font

    widths = [20, 18, 22, 16, 10, 10, 8, 40, 16, 26, 16]
    for col_idx, width in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(col_idx)].width = width
    ws.freeze_panes = 'A2'

    out = io.BytesIO()
    wb.save(out)
    return out.getvalue()


# ───────────────────────── PDF ─────────────────────────
def _build_pdf(qs, date_from, date_to):
    try:
        from reportlab.lib import colors
        from reportlab.lib.enums import TA_CENTER, TA_RIGHT
        from reportlab.lib.pagesizes import A4, landscape
        from reportlab.lib.styles import ParagraphStyle
        from reportlab.lib.units import mm
        from reportlab.platypus import (
            Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle,
        )
        # Reuse the shared font registration + RTL helpers from the survey PDF.
        from apps.surveys.pdf_report import _ensure_fonts, _fa_num, _rtl, _rtlp
    except ImportError as exc:  # pragma: no cover
        raise RuntimeError('خروجی PDF در دسترس نیست.') from exc

    _ensure_fonts()

    buf = io.BytesIO()
    page_size = landscape(A4)
    doc = SimpleDocTemplate(
        buf, pagesize=page_size,
        leftMargin=12 * mm, rightMargin=12 * mm, topMargin=12 * mm, bottomMargin=14 * mm,
        title='گزارش فعالیت‌ها',
    )
    ink = colors.HexColor('#1e293b')
    light = colors.HexColor('#e2e8f0')
    alt_row = colors.HexColor('#f8fafc')

    styles = {
        'title': ParagraphStyle('t', fontName='Vazir-Bold', fontSize=16, textColor=ink,
                                 alignment=TA_RIGHT, leading=22),
        'sub': ParagraphStyle('s', fontName='Vazir', fontSize=9.5,
                              textColor=colors.HexColor('#64748b'), alignment=TA_RIGHT, leading=15),
        'th': ParagraphStyle('th', fontName='Vazir-Bold', fontSize=8, textColor=colors.white,
                             alignment=TA_CENTER, leading=11),
        'cell': ParagraphStyle('c', fontName='Vazir', fontSize=7.5, textColor=ink,
                               alignment=TA_RIGHT, leading=10),
        'cellc': ParagraphStyle('cc', fontName='Vazir', fontSize=7.5, textColor=ink,
                                alignment=TA_CENTER, leading=10),
    }

    total = qs.count()
    story = [
        Paragraph(_rtlp('گزارش فعالیت‌ها و ممیزی سیستم'), styles['title']),
        Paragraph(
            _rtlp(
                f'بازه: {date_from:%Y-%m-%d} تا {date_to:%Y-%m-%d}  ·  '
                f'تعداد کل رکوردها: {_fa_num(total)}'
            ),
            styles['sub'],
        ),
        Spacer(1, 8),
    ]

    # Compact, PDF-friendly column subset.
    headers = ['زمان', 'نوع فعالیت', 'کاربر', 'وضعیت', 'شرح']
    data = [[Paragraph(_rtlp(h), styles['th']) for h in headers]]
    style_cmds = [
        ('BACKGROUND', (0, 0), (-1, 0), ink),
        ('GRID', (0, 0), (-1, -1), 0.4, light),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ('TOPPADDING', (0, 0), (-1, -1), 3),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 3),
        ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, alt_row]),
    ]

    rendered = 0
    for log in qs.iterator(chunk_size=500):
        if rendered >= PDF_MAX_ROWS:
            break
        failed = log.status == ActivityLog.STATUS_FAILED
        status_style = ParagraphStyle(
            'st', parent=styles['cellc'],
            textColor=colors.HexColor('#b91c1c') if failed else colors.HexColor('#16a34a'),
            fontName='Vazir-Bold',
        )
        data.append([
            Paragraph(_rtlp(_fmt_dt(log.created_at)), styles['cellc']),
            Paragraph(_rtlp(log.action_label), styles['cell']),
            Paragraph(_rtlp(log.actor_username or '—'), styles['cell']),
            Paragraph(_rtlp(_STATUS_LABELS.get(log.status, log.status)), status_style),
            Paragraph(_rtlp((log.description or '')[:120]), styles['cell']),
        ])
        if log.is_critical:
            style_cmds.append(('BACKGROUND', (0, len(data) - 1), (-1, len(data) - 1),
                               colors.HexColor('#fef2f2')))
        rendered += 1

    content_w = page_size[0] - 24 * mm
    table = Table(
        data,
        colWidths=[content_w * x for x in (0.16, 0.18, 0.18, 0.10, 0.38)],
        repeatRows=1,
    )
    table.setStyle(TableStyle(style_cmds))
    story.append(table)

    if total > rendered:
        story.append(Spacer(1, 8))
        story.append(Paragraph(
            _rtlp(
                f'+ {_fa_num(total - rendered)} رکورد دیگر در این بازه. '
                'برای فهرست کامل، خروجی Excel یا CSV را دانلود کنید.'
            ),
            styles['sub'],
        ))

    doc.build(story)
    return buf.getvalue()


def export_activity_logs(qs, export_format, date_from, date_to):
    """Return (content_bytes, content_type, filename) for the requested format."""
    filename = _filename(export_format, date_from, date_to)
    if export_format == 'csv':
        return _build_csv(qs), 'text/csv; charset=utf-8-sig', filename
    if export_format == 'excel':
        content = _build_excel(qs)
        return (
            content,
            'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            filename,
        )
    if export_format == 'pdf':
        return _build_pdf(qs, date_from, date_to), 'application/pdf', filename
    raise RuntimeError('فرمت خروجی نامعتبر است.')
